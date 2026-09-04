import os
import re
import json
import time
import hmac
import hashlib
import secrets
import sqlite3
import asyncio
from datetime import datetime
from pathlib import Path
import math
from urllib.parse import parse_qsl, unquote, quote
from urllib.request import Request, urlopen

from fastapi import FastAPI, HTTPException, Header, UploadFile, File
from fastapi.responses import FileResponse
from pydantic import BaseModel
import uvicorn
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.page import PageMargins
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from PIL import Image, ImageDraw, ImageFont

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, WebAppInfo
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters,
)

# =========================================================
# CONFIG
# =========================================================

BOT_TOKEN = os.getenv("BOT_TOKEN", "")
WEB_APP_URL = os.getenv("RENDER_EXTERNAL_URL", "").rstrip("/")
# Gemini is optional, but when configured it gives SERTAL a real-time Google Search
# grounding tool. Keep the key server-side in Render Environment Variables.
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "").strip() or os.getenv("GOOGLE_API_KEY", "").strip()
GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-3.8-flash").strip()

# Telegram IDs are used for bot-side admin permissions and message delivery.
ADMIN_IDS = {
    8357023784,
    7003441441,
}

# Separate personal Web App logins for administrators.
# Keep these credentials on the server; they are not exposed to the public UI.
ADMIN_ACCOUNTS = {
    "777": "администратор",
    "778": "администратор2",
    "779": "администратор3",
}

# Persistent database location. On Render, mount a Persistent Disk at /var/data.
# The env var allows the same code to run locally without changing the database.
_default_db_dir = Path("/var/data") if Path("/var/data").exists() else Path("data")
_default_db_dir.mkdir(parents=True, exist_ok=True)
DB_PATH = os.getenv("DB_PATH", str(_default_db_dir / "sertal_delivery.db"))
# If an older deploy used ./restaran.db, migrate it once into the persistent location.
_legacy_db = Path("restaran.db")
if not Path(DB_PATH).exists() and _legacy_db.exists() and Path(DB_PATH).resolve() != _legacy_db.resolve():
    try:
        import shutil
        Path(DB_PATH).parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(_legacy_db, DB_PATH)
        print(f"DB MIGRATED: {_legacy_db} -> {DB_PATH}")
    except Exception as exc:
        print("DB MIGRATION FAILED:", exc)
SUPPORT_GROUP_CHAT_ID = os.getenv("SUPPORT_GROUP_CHAT_ID", "").strip() or "-5588902525"
try:
    SUPPORT_GROUP_CHAT_ID = int(SUPPORT_GROUP_CHAT_ID) if SUPPORT_GROUP_CHAT_ID else None
except ValueError:
    SUPPORT_GROUP_CHAT_ID = None

EXPORT_DIR = Path("exports")
EXPORT_DIR.mkdir(exist_ok=True)
UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)
MAX_ORDER_WEIGHT_KG = 8.0
HEAVY_WEIGHT_KG = 4.0
HEAVY_ORDER_BONUS_AMD = 600
BATCH_MAX_ORDERS = 2
BATCH_MAX_DISTANCE_KM = 1.0

app = FastAPI()

# =========================================================
# DATABASE
# =========================================================

def db():
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=30000")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def get_support_group_chat_id():
    if SUPPORT_GROUP_CHAT_ID:
        return SUPPORT_GROUP_CHAT_ID
    try:
        conn=db()
        row=conn.execute("SELECT value FROM bot_settings WHERE key='support_group_chat_id'").fetchone()
        conn.close()
        return int(row["value"]) if row else None
    except Exception:
        return None


def set_support_group_chat_id(chat_id):
    conn=db()
    conn.execute("INSERT INTO bot_settings(key,value) VALUES('support_group_chat_id',?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (str(int(chat_id)),))
    conn.commit(); conn.close()


def init_db():
    conn = db()

    conn.execute("""
    CREATE TABLE IF NOT EXISTS users(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        phone TEXT UNIQUE NOT NULL,
        pin_hash TEXT NOT NULL,
        pin_plain TEXT,
        role TEXT NOT NULL DEFAULT 'customer',
        telegram_id INTEGER,
        created_at INTEGER NOT NULL,
        active INTEGER NOT NULL DEFAULT 1
    )
    """)

    conn.execute("""
    CREATE TABLE IF NOT EXISTS couriers(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        approved INTEGER NOT NULL DEFAULT 0,
        online INTEGER NOT NULL DEFAULT 0,
        lat REAL,
        lon REAL,
        updated_at INTEGER,
        active INTEGER NOT NULL DEFAULT 1
    )
    """)

    conn.execute("""
    CREATE TABLE IF NOT EXISTS restaurants(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        address TEXT NOT NULL DEFAULT '',
        phone TEXT NOT NULL DEFAULT '',
        active INTEGER NOT NULL DEFAULT 1,
        created_at INTEGER NOT NULL
    )
    """)

    conn.execute("""
    CREATE TABLE IF NOT EXISTS menu_items(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        restaurant_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        description TEXT NOT NULL DEFAULT '',
        price REAL NOT NULL DEFAULT 0,
        active INTEGER NOT NULL DEFAULT 1,
        created_at INTEGER NOT NULL,
        FOREIGN KEY(restaurant_id) REFERENCES restaurants(id) ON DELETE CASCADE
    )
    """)

    conn.execute("""
    CREATE TABLE IF NOT EXISTS orders(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        customer_id INTEGER NOT NULL,
        courier_id INTEGER,
        title TEXT NOT NULL,
        address TEXT NOT NULL,
        price REAL NOT NULL DEFAULT 0,
        status TEXT NOT NULL DEFAULT 'new',
        created_at INTEGER NOT NULL,
        customer_confirmed INTEGER NOT NULL DEFAULT 0,
        closed_at INTEGER,
        weight_kg REAL NOT NULL DEFAULT 0,
        courier_bonus_amd INTEGER NOT NULL DEFAULT 0,
        lat REAL,
        lon REAL,
        restaurant_address TEXT,
        restaurant_lat REAL,
        restaurant_lon REAL
    )
    """)

    conn.execute("""
    CREATE TABLE IF NOT EXISTS sessions(
        token TEXT PRIMARY KEY,
        user_id INTEGER NOT NULL,
        role TEXT NOT NULL,
        created_at INTEGER NOT NULL
    )
    """)

    conn.execute("""
    CREATE TABLE IF NOT EXISTS chat_messages(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        sender_role TEXT NOT NULL,
        text TEXT NOT NULL DEFAULT '',
        file_name TEXT,
        created_at INTEGER NOT NULL
    )
    """)

    conn.execute("""
    CREATE TABLE IF NOT EXISTS admin_message_bridge(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        admin_telegram_id INTEGER NOT NULL,
        telegram_message_id INTEGER NOT NULL,
        user_id INTEGER NOT NULL,
        created_at INTEGER NOT NULL,
        UNIQUE(admin_telegram_id, telegram_message_id)
    )
    """)

    conn.execute("""
    CREATE TABLE IF NOT EXISTS promos(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        code TEXT UNIQUE NOT NULL,
        kind TEXT NOT NULL DEFAULT 'percent',
        value REAL NOT NULL DEFAULT 0,
        min_order REAL NOT NULL DEFAULT 0,
        max_uses INTEGER,
        used_count INTEGER NOT NULL DEFAULT 0,
        active INTEGER NOT NULL DEFAULT 1,
        created_at INTEGER NOT NULL
    )
    """)

    conn.execute("""
    CREATE TABLE IF NOT EXISTS bot_settings(
        key TEXT PRIMARY KEY,
        value TEXT NOT NULL
    )
    """)

    # migrations
    migrations = [
        ("users", "pin_plain", "TEXT"),
        ("users", "telegram_id", "INTEGER"),
        ("users", "active", "INTEGER NOT NULL DEFAULT 1"),
        ("couriers", "approved", "INTEGER NOT NULL DEFAULT 0"),
        ("couriers", "online", "INTEGER NOT NULL DEFAULT 0"),
        ("couriers", "lat", "REAL"),
        ("couriers", "lon", "REAL"),
        ("couriers", "updated_at", "INTEGER"),
        ("couriers", "active", "INTEGER NOT NULL DEFAULT 1"),
        ("orders", "customer_confirmed", "INTEGER NOT NULL DEFAULT 0"),
        ("orders", "closed_at", "INTEGER"),
        ("orders", "weight_kg", "REAL NOT NULL DEFAULT 0"),
        ("orders", "courier_bonus_amd", "INTEGER NOT NULL DEFAULT 0"),
        ("orders", "lat", "REAL"),
        ("orders", "lon", "REAL"),
        ("orders", "restaurant_address", "TEXT"),
        ("orders", "restaurant_lat", "REAL"),
        ("orders", "restaurant_lon", "REAL"),
        ("orders", "restaurant_name", "TEXT"),
        ("orders", "delivery_deadline", "TEXT"),
        ("orders", "delivered_at", "TEXT"),
        ("orders", "floor", "TEXT"),
        ("orders", "entrance", "TEXT"),
        ("orders", "apartment", "TEXT"),
        ("orders", "intercom", "TEXT"),
        ("orders", "recipient_name", "TEXT"),
        ("orders", "delivery_note", "TEXT"),
        ("orders", "payment_note", "TEXT"),
        ("orders", "items_text", "TEXT"),
        ("orders", "change_amount", "REAL"),
        ("orders", "route_no", "INTEGER"),
        ("orders", "restaurant_id", "INTEGER"),
        ("orders", "payment_method", "TEXT NOT NULL DEFAULT 'cash'"),
        ("orders", "payment_status", "TEXT NOT NULL DEFAULT 'cash'"),
        ("orders", "receipt_file", "TEXT"),
        ("orders", "promo_code", "TEXT"),
        ("orders", "discount_amd", "REAL NOT NULL DEFAULT 0"),
        ("users", "default_address", "TEXT"),
        ("users", "default_lat", "REAL"),
        ("users", "default_lon", "REAL"),
    ]

    for table, column, definition in migrations:
        cols = [
            r["name"]
            for r in conn.execute(
                f"PRAGMA table_info({table})"
            ).fetchall()
        ]

        if column not in cols:
            conn.execute(
                f"ALTER TABLE {table} ADD COLUMN {column} {definition}"
            )

    conn.commit()
    conn.close()


init_db()

# =========================================================
# HELPERS
# =========================================================

def normalize_phone(phone):
    return re.sub(r"\D", "", str(phone or ""))


def phone_digits(phone):
    return re.sub(r"\D", "", str(phone or ""))


def phones_equal(a, b):
    a = phone_digits(a)
    b = phone_digits(b)

    if not a or not b:
        return False

    return (
        a == b
        or a.lstrip("0") == b.lstrip("0")
        or a.endswith(b)
        or b.endswith(a)
    )


def hash_pin(pin):
    return hashlib.sha256(str(pin).encode()).hexdigest()


def check_pin(pin, password_hash):
    return hmac.compare_digest(
        hash_pin(pin),
        password_hash
    )


def new_pin():
    return str(secrets.randbelow(900000) + 100000)


def create_session(user_id, role):
    token = secrets.token_urlsafe(48)

    conn = db()
    conn.execute(
        "INSERT INTO sessions(token,user_id,role,created_at) VALUES(?,?,?,?)",
        (token, user_id, role, int(time.time()))
    )
    conn.commit()
    conn.close()

    return token


def get_session(authorization):
    if not authorization:
        return None

    token = authorization.replace("Bearer ", "").strip()

    if not token:
        return None

    conn = db()

    row = conn.execute("""
        SELECT
            s.token,
            s.user_id,
            s.role,
            u.name,
            u.phone,
            u.telegram_id,
            u.active
        FROM sessions s
        JOIN users u ON u.id=s.user_id
        WHERE s.token=?
    """, (token,)).fetchone()

    conn.close()

    if not row or not row["active"]:
        return None

    return row


def require_user(authorization):
    user = get_session(authorization)

    if not user:
        raise HTTPException(
            status_code=401,
            detail="Требуется авторизация"
        )

    return user


def require_admin(authorization):
    user = require_user(authorization)

    if user["role"] != "admin":
        raise HTTPException(
            status_code=403,
            detail="Нет доступа"
        )

    return user


def cleanup_old_closed():
    # История заказов/логи не удаляются автоматически.
    return


def order_cutoff():
    return int(time.time()) - 300


def haversine_km(lat1, lon1, lat2, lon2):
    """Distance between two GPS points in kilometres."""
    r = 6371.0088
    p1 = math.radians(float(lat1))
    p2 = math.radians(float(lat2))
    dp = math.radians(float(lat2) - float(lat1))
    dl = math.radians(float(lon2) - float(lon1))
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def geocode_yerevan(address):
    """Best-effort geocoding. Orders stay valid even if geocoding is unavailable."""
    try:
        query = str(address or "").strip()
        if not query:
            return None, None
        if "ереван" not in query.lower() and "yerevan" not in query.lower():
            query = f"{query}, Yerevan, Armenia"
        url = "https://nominatim.openstreetmap.org/search?" +               f"format=jsonv2&limit=1&countrycodes=am&q={quote(query)}"
        req = Request(url, headers={"User-Agent": "SERTAL DELIVERY/1.0 order-map"})
        with urlopen(req, timeout=5) as response:
            data = json.loads(response.read().decode("utf-8"))
        if data:
            return float(data[0]["lat"]), float(data[0]["lon"])
    except Exception as e:
        print("GEOCODE:", e)
    return None, None


def calculate_courier_bonus(weight_kg):
    weight = float(weight_kg or 0)
    return HEAVY_ORDER_BONUS_AMD if weight > HEAVY_WEIGHT_KG else 0


def courier_active_order_count(conn, courier_id):
    return conn.execute("""
        SELECT COUNT(*)
        FROM orders
        WHERE courier_id=?
        AND status IN ('assigned','accepted','delivering')
    """, (courier_id,)).fetchone()[0]


def batch_distance_to_existing(conn, courier_id, new_order):
    """Returns the shortest known delivery-point distance in km."""
    if new_order["lat"] is None or new_order["lon"] is None:
        return None
    rows = conn.execute("""
        SELECT lat, lon
        FROM orders
        WHERE courier_id=?
        AND status IN ('assigned','accepted','delivering')
        AND lat IS NOT NULL AND lon IS NOT NULL
    """, (courier_id,)).fetchall()
    if not rows:
        return None
    return min(
        haversine_km(new_order["lat"], new_order["lon"], r["lat"], r["lon"])
        for r in rows
    )


# =========================================================
# TELEGRAM WEBAPP VALIDATION
# =========================================================

def validate_telegram_init_data(init_data):
    if not BOT_TOKEN or not init_data:
        return None

    try:
        pairs = dict(parse_qsl(init_data, keep_blank_values=True))

        received_hash = pairs.pop("hash", None)

        if not received_hash:
            return None

        data_check_string = "\n".join(
            f"{key}={value}"
            for key, value in sorted(pairs.items())
        )

        secret_key = hmac.new(
            b"WebAppData",
            BOT_TOKEN.encode(),
            hashlib.sha256
        ).digest()

        calculated_hash = hmac.new(
            secret_key,
            data_check_string.encode(),
            hashlib.sha256
        ).hexdigest()

        if not hmac.compare_digest(
            calculated_hash,
            received_hash
        ):
            return None

        auth_date = int(pairs.get("auth_date", "0"))

        if int(time.time()) - auth_date > 86400:
            return None

        user_raw = pairs.get("user")

        if not user_raw:
            return None

        user = json.loads(unquote(user_raw))

        return user

    except Exception as e:
        print("Telegram validation:", e)
        return None


# =========================================================
# RECEIPTS / RESTAURANT HELPERS
# =========================================================


def make_order_receipt_image(order_id):
    """Create a premium SERTAL receipt image for a real order."""
    conn = db()
    row = conn.execute("""
        SELECT o.*, u.name AS customer_name, u.phone AS customer_phone,
               cu.name AS courier_name, cu.phone AS courier_phone
        FROM orders o
        JOIN users u ON u.id=o.customer_id
        LEFT JOIN couriers c ON c.id=o.courier_id
        LEFT JOIN users cu ON cu.id=c.user_id
        WHERE o.id=?
    """, (order_id,)).fetchone()
    conn.close()
    if not row:
        return None

    W, H = 1080, 1480
    bg=(8,8,8); card=(18,18,18); gold=(212,175,55); cream=(245,241,230); muted=(160,152,135); line=(54,47,30)
    img=Image.new("RGB", (W,H), bg)
    d=ImageDraw.Draw(img)
    d.rounded_rectangle((35,35,W-35,H-35), radius=42, fill=card, outline=gold, width=3)

    # SERTAL emblem: a clean vector mark, no external image dependency.
    d.rounded_rectangle((72,68,190,186), radius=28, outline=gold, width=5)
    d.line((105,128,157,128), fill=gold, width=8)
    d.line((131,101,131,155), fill=gold, width=8)
    d.arc((91,91,171,171), 210, 330, fill=cream, width=4)
    d.text((220,72), "SERTAL", font=_font(48,True), fill=gold)
    d.text((222,130), "DELIVERY", font=_font(27,True), fill=cream)
    d.text((W-315,78), "RECEIPT", font=_font(22,True), fill=muted)
    d.text((W-315,112), f"#{order_id}", font=_font(34,True), fill=cream)

    d.line((72,215,W-72,215), fill=line, width=3)
    d.text((72,248), "ЧЕК ЗАКАЗА", font=_font(34,True), fill=cream)
    d.text((72,292), datetime.fromtimestamp(row['created_at']).strftime('%d.%m.%Y  ·  %H:%M'), font=_font(23), fill=muted)

    y=355
    sections=[
        ("КЛИЕНТ", str(row['customer_name'] or '—')),
        ("ТЕЛЕФОН", "+"+str(row['customer_phone'] or '').lstrip('+') if row['customer_phone'] else '—'),
        ("АДРЕС ДОСТАВКИ", str(row['address'] or '—')),
        ("ПОЛУЧАТЕЛЬ", str(row['recipient_name'] or '—') if 'recipient_name' in row.keys() else '—'),
    ]
    for label,value in sections:
        d.text((72,y), label, font=_font(17,True), fill=gold)
        # wrap long values
        words=value.split(); lines=[]; cur=''
        for word in words:
            test=(cur+' '+word).strip()
            if len(test)>54:
                lines.append(cur); cur=word
            else: cur=test
        if cur: lines.append(cur)
        for ln in lines[:2]:
            y+=29; d.text((72,y),ln,font=_font(24),fill=cream)
        y+=43

    d.line((72,y,W-72,y),fill=line,width=2); y+=32
    d.text((72,y), "СОСТАВ ЗАКАЗА", font=_font(18,True), fill=gold); y+=42
    items = str(row['items_text'] or row['title'] or 'Заказ').splitlines() if 'items_text' in row.keys() else [str(row['title'] or 'Заказ')]
    for item in items[:8]:
        item=item.strip()
        if not item: continue
        d.ellipse((74,y+7,84,y+17),fill=gold)
        d.text((102,y),item[:70],font=_font(23),fill=cream); y+=37

    y+=10; d.line((72,y,W-72,y),fill=line,width=2); y+=30
    payment = row['payment_method'] if 'payment_method' in row.keys() else 'cash'
    payment_text = 'НАЛИЧНЫЕ' if payment=='cash' else 'ОНЛАЙН'
    d.text((72,y), "ОПЛАТА", font=_font(18,True), fill=gold)
    d.text((260,y), payment_text, font=_font(23,True), fill=cream)
    y+=42
    if 'change_amount' in row.keys() and float(row['change_amount'] or 0)>0:
        d.text((72,y), "СДАЧА", font=_font(18,True), fill=gold)
        d.text((260,y), f"{float(row['change_amount']):,.0f}".replace(',',' ') + " AMD", font=_font(23,True), fill=cream)
        y+=42

    total=float(row['price'] or 0)
    d.rounded_rectangle((650,y-20,1008,y+92),radius=22,outline=gold,width=2)
    d.text((682,y),"ИТОГО",font=_font(20,True),fill=gold)
    d.text((682,y+35),f"{total:,.0f}".replace(',',' ') + " AMD",font=_font(31,True),fill=cream)
    y+=135

    status_names={"new":"НОВЫЙ","assigned":"НАЗНАЧЕН","accepted":"ПРИНЯТ","delivering":"В ДОСТАВКЕ","delivered":"ДОСТАВЛЕН","closed":"ЗАКРЫТ"}
    d.text((72,y),"СТАТУС",font=_font(18,True),fill=gold)
    d.text((220,y),status_names.get(str(row['status']),str(row['status']).upper()),font=_font(23,True),fill=cream)
    y+=58
    d.line((72,y,W-72,y),fill=line,width=2); y+=35
    d.text((72,y),"SERTAL DELIVERY",font=_font(22,True),fill=gold)
    d.text((72,y+34),"Спасибо за заказ. Чек сформирован автоматически.",font=_font(19),fill=muted)
    d.text((W-360,y+34),"AMD · ARMENIA",font=_font(17,True),fill=muted)

    path=EXPORT_DIR/f"sertal_receipt_{order_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
    img.save(path,"PNG",optimize=True)
    return path

def generate_receipt(order_id):
    conn = db()
    row = conn.execute("""
        SELECT o.*, u.name AS customer_name, u.phone AS customer_phone
        FROM orders o JOIN users u ON u.id=o.customer_id WHERE o.id=?
    """, (order_id,)).fetchone()
    conn.close()
    if not row:
        return None
    path = EXPORT_DIR / f"sertal_order_{order_id}.pdf"
    c = canvas.Canvas(str(path), pagesize=A4)
    w, h = A4
    y = h - 55
    c.setFont("Helvetica-Bold", 22); c.drawString(55, y, "SERTAL DELIVERY")
    y -= 38
    c.setFont("Helvetica-Bold", 16); c.drawString(55, y, f"Заказ №{order_id}")
    y -= 28
    c.setFont("Helvetica", 11)
    payment = row["payment_method"] if "payment_method" in row.keys() else "cash"
    lines = [
        f"Клиент: {row['customer_name']}",
        f"Телефон: +{row['customer_phone']}",
        f"Адрес доставки: {row['address']}",
        f"Что доставить: {row['title']}",
        f"Ресторан: {row['restaurant_name'] or '—'}",
        f"Способ оплаты: {'Наличные' if payment == 'cash' else 'Онлайн'}",
        f"Сумма: {float(row['price'] or 0):.0f} ֏",
        f"Статус: {row['status']}",
        f"Создан: {datetime.fromtimestamp(row['created_at']).strftime('%d.%m.%Y %H:%M')}",
    ]
    for line in lines:
        c.drawString(55, y, line[:110]); y -= 22
    c.line(55, y, w-55, y); y -= 28
    c.setFont("Helvetica-Oblique", 10); c.drawString(55, y, "Документ сформирован автоматически SERTAL DELIVERY")
    c.save()
    conn = db(); conn.execute("UPDATE orders SET receipt_file=? WHERE id=?", (str(path), order_id)); conn.commit(); conn.close()
    return path

async def send_receipt_to_customer(user, path, order_id):
    if not path or not telegram_app or not user["telegram_id"]:
        return
    try:
        with open(path, "rb") as fh:
            await telegram_app.bot.send_document(user["telegram_id"], fh, filename=path.name, caption=f"SERTAL DELIVERY — чек заказа №{order_id}")
    except Exception as exc:
        print("RECEIPT TELEGRAM:", exc)

# =========================================================
# MODELS
# =========================================================

class LoginData(BaseModel):
    phone: str
    telegram_id: int | None = None




class CustomerCreate(BaseModel):
    name: str
    phone: str


class CourierCreate(BaseModel):
    name: str
    phone: str


class OrderCreate(BaseModel):
    phone: str
    title: str
    address: str
    restaurant_address: str = ""
    restaurant_name: str = ""
    price: float = 0
    weight_kg: float = 0
    delivery_deadline: str = ""
    floor: str = ""
    entrance: str = ""
    apartment: str = ""
    intercom: str = ""
    recipient_name: str = ""
    delivery_note: str = ""
    payment_note: str = ""
    items_text: str = ""
    change_amount: float = 0


class AssignData(BaseModel):
    courier_id: int


class OnlineData(BaseModel):
    online: bool


class LocationData(BaseModel):
    lat: float
    lon: float


class ChatMessageData(BaseModel):
    text: str = ""

class RestaurantCreate(BaseModel):
    name: str
    address: str = ""
    phone: str = ""

class MenuItemCreate(BaseModel):
    restaurant_id: int
    name: str
    description: str = ""
    price: float

class CartItem(BaseModel):
    menu_item_id: int
    quantity: int = 1

class CartOrderData(BaseModel):
    restaurant_id: int
    items: list[CartItem]
    address: str
    payment_method: str = "cash"
    delivery_note: str = ""
    promo_code: str = ""

class CustomerAddressData(BaseModel):
    address: str = ""
    lat: float | None = None
    lon: float | None = None


# =========================================================
# WEB
# =========================================================

@app.get("/")
async def index():
    return FileResponse(Path(__file__).resolve().with_name("app.html"))


@app.head("/")
async def head_index():
    return {}


# =========================================================
# LOGIN
# =========================================================

@app.post("/api/login")
async def login(data: LoginData):
    # Обычный вход покупателя/курьера.
    # Специальные номера администраторов (+777/+778/+779)
    # разрешены отдельным персональным входом, но не считаются
    # обычными телефонными номерами.
    phone = normalize_phone(data.phone)
    if phone in ADMIN_ACCOUNTS:
        raise HTTPException(
            status_code=403,
            detail="Это номер администратора. Нажмите «Админ» и введите персональный никнейм."
        )
    # Публичная регистрация включена. Для реальных номеров разрешаем
    # международный формат после очистки от пробелов, скобок и дефисов.
    # Минимум 3 цифры оставляем только для тестовых номеров администраторов.
    if len(phone) < 3:
        raise HTTPException(status_code=400, detail="Введите корректный номер телефона")

    conn = db()
    user = conn.execute("""
        SELECT *
        FROM users
        WHERE role IN ('customer','courier')
          AND active=1
          AND phone=?
    """, (phone,)).fetchone()

    if not user:
        conn.close()
        raise HTTPException(status_code=401, detail="Этот номер не добавлен администратором")

    if user["role"] == "courier":
        courier = conn.execute("SELECT * FROM couriers WHERE user_id=?", (user["id"],)).fetchone()
        if not courier or not courier["approved"]:
            conn.close()
            raise HTTPException(status_code=403, detail="Курьер ещё не одобрен администратором")
        if not courier["active"]:
            conn.close()
            raise HTTPException(status_code=403, detail="Курьер деактивирован")

    if data.telegram_id:
        conn.execute("UPDATE users SET telegram_id=? WHERE id=?", (int(data.telegram_id), user["id"]))
        conn.commit()
    token = create_session(user["id"], user["role"])
    conn.close()
    return {"token": token, "role": user["role"]}


# =========================================================
# ME
# =========================================================

@app.get("/api/me")
async def me(
    authorization: str = Header(default="")
):
    user = require_user(authorization)

    return {
        "id": user["user_id"],
        "name": user["name"],
        "phone": user["phone"],
        "telegram_id": user["telegram_id"],
        "role": user["role"]
    }


# =========================================================
# ADMIN STATS
# =========================================================

@app.get("/api/admin/stats")
async def admin_stats(
    authorization: str = Header(default="")
):
    require_admin(authorization)
    cleanup_old_closed()

    conn = db()

    customers = conn.execute("""
        SELECT COUNT(*)
        FROM users
        WHERE role='customer'
        AND active=1
    """).fetchone()[0]

    couriers = conn.execute("""
        SELECT COUNT(*)
        FROM couriers c
        JOIN users u ON u.id=c.user_id
        WHERE c.active=1
        AND u.active=1
    """).fetchone()[0]

    active_orders = conn.execute("""
        SELECT COUNT(*)
        FROM orders
        WHERE status!='closed'
    """).fetchone()[0]

    delivered = conn.execute("""
        SELECT COUNT(*)
        FROM orders
        WHERE status='delivered'
    """).fetchone()[0]

    closed = conn.execute("""
        SELECT COUNT(*)
        FROM orders
        WHERE status='closed'
    """).fetchone()[0]

    revenue = conn.execute("""
        SELECT COALESCE(SUM(price),0)
        FROM orders
        WHERE status IN ('delivered','closed')
    """).fetchone()[0]

    conn.close()

    return {
        "customers": customers,
        "couriers": couriers,
        "active_orders": active_orders,
        "delivered": delivered,
        "closed": closed,
        "revenue": revenue
    }


# =========================================================
# CUSTOMERS
# =========================================================

@app.get("/api/admin/customers")
async def admin_customers(
    authorization: str = Header(default="")
):
    require_admin(authorization)

    conn = db()

    rows = conn.execute("""
        SELECT id,name,phone,active,created_at
        FROM users
        WHERE role='customer'
        ORDER BY active DESC, id DESC
    """).fetchall()

    conn.close()

    return [dict(x) for x in rows]


@app.post("/api/admin/customers")
async def admin_create_customer(
    data: CustomerCreate,
    authorization: str = Header(default="")
):
    require_admin(authorization)

    phone = normalize_phone(data.phone)

    if len(phone) < 5:
        raise HTTPException(
            status_code=400,
            detail="Неверный номер телефона"
        )

    conn = db()

    try:
        cur = conn.execute("""
            INSERT INTO users(
                name,
                phone,
                pin_hash,
                pin_plain,
                role,
                created_at,
                active
            )
            VALUES(?,?,?,?,?,?,1)
        """, (
            data.name.strip(),
            phone,
            "",
            "",
            "customer",
            int(time.time())
        ))

        conn.commit()

        return {
            "id": cur.lastrowid,
            "name": data.name.strip(),
            "phone": phone,
        }

    except sqlite3.IntegrityError:
        raise HTTPException(
            status_code=400,
            detail="Клиент с таким номером уже существует"
        )

    finally:
        conn.close()


# =========================================================
# CUSTOMER DEACTIVATION — SOFT DELETE
# =========================================================

@app.post("/api/admin/users/{user_id}/delete")
async def admin_delete_user(user_id: int, authorization: str = Header(default="")):
    require_admin(authorization)
    conn=db()
    try:
        user=conn.execute("SELECT id,role FROM users WHERE id=?",(user_id,)).fetchone()
        if not user or user["role"]=="admin":
            raise HTTPException(status_code=404,detail="Участник не найден")
        # Full deletion: remove the participant, courier record, sessions,
        # support chat and orders owned by this participant.
        conn.execute("DELETE FROM admin_message_bridge WHERE user_id=?",(user_id,))
        conn.execute("DELETE FROM chat_messages WHERE user_id=?",(user_id,))
        conn.execute("DELETE FROM sessions WHERE user_id=?",(user_id,))
        courier=conn.execute("SELECT id FROM couriers WHERE user_id=?",(user_id,)).fetchone()
        if courier:
            conn.execute("DELETE FROM orders WHERE courier_id=?",(courier["id"],))
            conn.execute("DELETE FROM couriers WHERE id=?",(courier["id"],))
        conn.execute("DELETE FROM orders WHERE customer_id=?",(user_id,))
        conn.execute("DELETE FROM users WHERE id=?",(user_id,))
        conn.commit()
        return {"ok":True,"deleted_user_id":user_id}
    finally:
        conn.close()

@app.post("/api/admin/customers/{customer_id}/delete")
async def admin_delete_customer(customer_id: int, authorization: str = Header(default="")):
    return await admin_delete_user(customer_id, authorization)


# =========================================================
# COURIERS
# =========================================================

@app.get("/api/admin/couriers")
async def admin_couriers(
    authorization: str = Header(default="")
):
    require_admin(authorization)

    conn = db()

    rows = conn.execute("""
        SELECT
            c.id,
            c.user_id,
            c.approved,
            c.online,
            c.lat,
            c.lon,
            c.updated_at,
            c.active,
            u.name,
            u.phone
        FROM couriers c
        JOIN users u ON u.id=c.user_id
        ORDER BY c.id DESC
    """).fetchall()

    conn.close()

    return [dict(x) for x in rows]


@app.post("/api/admin/couriers")
async def admin_create_courier(
    data: CourierCreate,
    authorization: str = Header(default="")
):
    require_admin(authorization)

    phone = normalize_phone(data.phone)
    if len(phone) < 5:
        raise HTTPException(status_code=400, detail="Неверный номер телефона")
    conn = db()

    try:

        cur = conn.execute("""
            INSERT INTO users(
                name,
                phone,
                pin_hash,
                pin_plain,
                role,
                created_at,
                active
            )
            VALUES(?,?,?,?,?,?,1)
        """, (
            data.name.strip(),
            phone,
            "",
            "",
            "courier",
            int(time.time())
        ))

        user_id = cur.lastrowid

        conn.execute("""
            INSERT INTO couriers(
                user_id,
                approved,
                online,
                active
            )
            VALUES(?,0,0,1)
        """, (user_id,))

        conn.commit()

        return {
            "id": user_id,
            "name": data.name.strip(),
            "phone": phone,
        }

    except sqlite3.IntegrityError:
        raise HTTPException(
            status_code=400,
            detail="Курьер с таким номером уже существует"
        )

    finally:
        conn.close()


@app.post("/api/admin/couriers/{courier_id}/approve")
async def approve_courier(
    courier_id: int,
    authorization: str = Header(default="")
):
    require_admin(authorization)

    conn = db()

    conn.execute("""
        UPDATE couriers
        SET approved=1, active=1
        WHERE id=?
    """, (courier_id,))

    conn.execute("""
        UPDATE users
        SET active=1
        WHERE id=(
            SELECT user_id
            FROM couriers
            WHERE id=?
        )
    """, (courier_id,))

    conn.commit()
    conn.close()

    return {"ok": True}


@app.post("/api/admin/couriers/{courier_id}/fire")
async def fire_courier(
    courier_id: int,
    authorization: str = Header(default="")
):
    require_admin(authorization)

    conn = db()

    courier = conn.execute(
        "SELECT user_id FROM couriers WHERE id=?",
        (courier_id,)
    ).fetchone()

    if not courier:
        conn.close()
        raise HTTPException(
            status_code=404,
            detail="Курьер не найден"
        )

    conn.execute("""
        UPDATE couriers
        SET active=0, approved=0, online=0
        WHERE id=?
    """, (courier_id,))

    conn.execute("""
        UPDATE users
        SET active=0
        WHERE id=?
    """, (courier["user_id"],))

    # Снимаем активные заказы
    conn.execute("""
        UPDATE orders
        SET courier_id=NULL,
            status='new'
        WHERE courier_id=?
        AND status IN ('assigned','accepted','delivering')
    """, (courier_id,))

    conn.commit()
    conn.close()

    return {"ok": True}


# =========================================================
# ORDERS
# =========================================================

@app.get("/api/admin/map-orders")
async def admin_map_orders(
    authorization: str = Header(default="")
):
    require_admin(authorization)
    cleanup_old_closed()
    conn = db()
    rows = conn.execute("""
        SELECT o.id, o.title, o.address, o.price, o.status,
               o.weight_kg, o.courier_bonus_amd, o.lat, o.lon,
               o.restaurant_address, o.restaurant_lat, o.restaurant_lon,
               u.name AS customer_name,
               co.name AS courier_name
        FROM orders o
        JOIN users u ON u.id=o.customer_id
        LEFT JOIN couriers c ON c.id=o.courier_id
        LEFT JOIN users co ON co.id=c.user_id
        WHERE o.status!='closed'
        ORDER BY o.id DESC
    """).fetchall()
    couriers = conn.execute("""
        SELECT c.id, c.lat, c.lon, c.online, u.name
        FROM couriers c JOIN users u ON u.id=c.user_id
        WHERE c.active=1 AND c.lat IS NOT NULL AND c.lon IS NOT NULL
    """).fetchall()
    conn.close()
    return {"orders": [dict(x) for x in rows], "couriers": [dict(x) for x in couriers]}


@app.post("/api/admin/ai")
async def admin_ai(
    payload: dict,
    authorization: str = Header(default="")
):
    require_admin(authorization)
    question = str(payload.get("message", "")).strip()
    if not question:
        raise HTTPException(status_code=400, detail="Введите вопрос")

    # Natural-language admin actions: these are executed server-side only after
    # require_admin(), so the AI cannot bypass the existing admin permission.
    q = question.lower()
    phone_match = re.search(r"(\+?\d[\d\s()\-]{5,})$", question)
    if phone_match and any(x in q for x in ("добавь клиента", "создай клиента", "добавить клиента")):
        phone = normalize_phone(phone_match.group(1))
        name = question[:phone_match.start()].strip()
        name = re.sub(r"^(добавь|создай|добавить)\s+(клиента)\s*", "", name, flags=re.I).strip(" :,-")
        if not name or len(phone) < 5:
            raise HTTPException(status_code=400, detail="Укажите имя и номер клиента")
        pin = new_pin()
        conn = db()
        try:
            cur = conn.execute("""
                INSERT INTO users(name,phone,pin_hash,pin_plain,role,created_at,active)
                VALUES(?,?,?,?,?,?,1)
            """, (name, phone, hash_pin(pin), pin, "customer", int(time.time())))
            conn.commit()
            return {"answer": f"✅ Клиент «{name}» создан. ID: {cur.lastrowid}. PIN: {pin}", "mode": "action"}
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=400, detail="Клиент с таким номером уже существует")
        finally:
            conn.close()

    if phone_match and any(x in q for x in ("добавь курьера", "создай курьера", "добавить курьера")):
        phone = normalize_phone(phone_match.group(1))
        name = question[:phone_match.start()].strip()
        name = re.sub(r"^(добавь|создай|добавить)\s+(курьера)\s*", "", name, flags=re.I).strip(" :,-")
        if not name or len(phone) < 5:
            raise HTTPException(status_code=400, detail="Укажите имя и номер курьера")
        pin = new_pin()
        conn = db()
        try:
            cur = conn.execute("""
                INSERT INTO users(name,phone,pin_hash,pin_plain,role,created_at,active)
                VALUES(?,?,?,?,?,?,1)
            """, (name, phone, hash_pin(pin), pin, "courier", int(time.time())))
            user_id = cur.lastrowid
            conn.execute("INSERT INTO couriers(user_id,approved,online,active) VALUES(?,0,0,1)", (user_id,))
            conn.commit()
            return {"answer": f"✅ Курьер «{name}» создан. ID: {user_id}. PIN: {pin}. Теперь его можно одобрить в разделе «Курьеры».", "mode": "action"}
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=400, detail="Курьер с таким номером уже существует")
        finally:
            conn.close()

    # Optional real AI integration. If OPENAI_API_KEY is not configured,
    # the assistant still works in local operational mode.
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    model = os.getenv("OPENAI_MODEL", "gpt-5.6-luna")
    if api_key:
        try:
            conn = db()
            summary = conn.execute("""
                SELECT
                    COUNT(*) AS total,
                    SUM(CASE WHEN status='new' THEN 1 ELSE 0 END) AS new_count,
                    SUM(CASE WHEN status IN ('assigned','accepted','delivering') THEN 1 ELSE 0 END) AS active_count,
                    SUM(CASE WHEN status='delivered' THEN 1 ELSE 0 END) AS delivered_count,
                    COALESCE(SUM(courier_bonus_amd),0) AS heavy_bonus
                FROM orders
                WHERE status!='closed'
            """).fetchone()
            conn.close()

            prompt = (
                "Ты ИИ-помощник админ-панели SERTAL DELIVERY в Ереване. "
                "Отвечай на русском, структурированно и по делу. Учитывай текущие заказы, клиентов, курьеров, рестораны и поддержку. Не выдумывай данные. Если действие нельзя выполнить через доступные команды, честно скажи об этом. "
                f"Текущая статистика: {dict(summary)}. "
                f"Вопрос администратора: {question}"
            )
            body = json.dumps({
                "model": model,
                "input": prompt
            }).encode("utf-8")
            req = Request(
                "https://api.openai.com/v1/responses",
                data=body,
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {api_key}"
                },
                method="POST"
            )
            with urlopen(req, timeout=25) as response:
                result = json.loads(response.read().decode("utf-8"))
            answer = result.get("output_text")
            if not answer:
                parts = []
                for item in result.get("output", []):
                    for content in item.get("content", []):
                        if content.get("type") == "output_text":
                            parts.append(content.get("text", ""))
                answer = "\n".join(parts).strip()
            if answer:
                return {"answer": answer, "mode": "ai"}
        except Exception as e:
            print("AI:", e)

    # Local fallback: useful without an external API key.
    q = question.lower()
    conn = db()
    stats = conn.execute("""
        SELECT
            COUNT(*) AS total,
            SUM(CASE WHEN status='new' THEN 1 ELSE 0 END) AS new_count,
            SUM(CASE WHEN status IN ('assigned','accepted','delivering') THEN 1 ELSE 0 END) AS active_count,
            SUM(CASE WHEN status='delivered' THEN 1 ELSE 0 END) AS delivered_count,
            COALESCE(SUM(courier_bonus_amd),0) AS heavy_bonus
        FROM orders WHERE status!='closed'
    """).fetchone()
    conn.close()

    if any(x in q for x in ("клиент", "покупател", "зарегистрирован")):
        rows=conn.execute("SELECT name,phone,role,active FROM users WHERE role IN ('customer','courier') ORDER BY id DESC LIMIT 20").fetchall()
        answer="Клиенты и курьеры в базе:\n" + ("\n".join(f"{r['name']} — {r['phone']} — {r['role']} — {'активен' if r['active'] else 'неактивен'}" for r in rows) if rows else "База пуста.")
        conn.close()
        return {"answer":answer,"mode":"local"}

    if any(x in q for x in ("курьер", "курьеры")):
        rows=conn.execute("SELECT u.name,u.phone,c.approved,c.online,c.active FROM couriers c JOIN users u ON u.id=c.user_id ORDER BY c.id DESC").fetchall()
        answer="Курьеры:\n" + ("\n".join(f"{r['name']} — {r['phone']} — {'одобрен' if r['approved'] else 'не одобрен'} — {'онлайн' if r['online'] else 'офлайн'}" for r in rows) if rows else "Курьеров нет.")
        conn.close()
        return {"answer":answer,"mode":"local"}

    if any(x in q for x in ("ресторан", "рестораны", "меню")):
        try:
            rows=conn.execute("SELECT name,address,phone,active FROM restaurants ORDER BY id DESC").fetchall()
            answer="Рестораны:\n" + ("\n".join(f"{r['name']} — {r['address']} — {'активен' if r['active'] else 'выключен'}" for r in rows) if rows else "Рестораны ещё не добавлены.")
        except sqlite3.OperationalError:
            answer="Таблица ресторанов ещё не создана."
        conn.close()
        return {"answer":answer,"mode":"local"}

    if any(x in q for x in ("помощ", "что умеешь", "команд")):
        conn.close()
        return {"answer":"Я могу показать статистику, найти клиентов и курьеров, проверить рестораны, подсказать по активным заказам и выполнять разрешённые административные действия. Для выгрузки клиентской базы в Telegram используйте !logsfile.","mode":"local"}

    if any(x in q for x in ("статист", "заказ", "сколько", "заказы")):
        answer = (
            f"Всего активных заказов: {stats['total'] or 0}\n"
            f"Новых: {stats['new_count'] or 0}\n"
            f"В работе: {stats['active_count'] or 0}\n"
            f"Доставлено: {stats['delivered_count'] or 0}\n"
            f"⚖️ Доплат за тяжёлые заказы: {stats['heavy_bonus'] or 0} AMD"
        )
    elif "правил" in q or "вес" in q or "4 кг" in q or "8 кг" in q:
        answer = "Правила: до 4 кг — обычная оплата; больше 4 до 8 кг — +600 AMD курьеру; больше 8 кг — заказ не создаётся."
    elif "2" in q or "килом" in q or "км" in q:
        answer = "Курьер может иметь максимум 2 активных заказа. Второй допускается, если известные точки выполнения находятся не дальше 1 км."
    else:
        answer = "Я могу помочь со статистикой заказов, правилами веса, доплатами и проверкой условий для второго заказа."

    return {"answer": answer, "mode": "local"}


@app.get("/api/admin/orders")
async def admin_orders(
    authorization: str = Header(default="")
):
    require_admin(authorization)
    cleanup_old_closed()

    cutoff = order_cutoff()

    conn = db()

    rows = conn.execute("""
        SELECT
            o.*,
            cu.name AS customer_name,
            cu.phone AS customer_phone,
            co.name AS courier_name,
            c.lat AS courier_lat,
            c.lon AS courier_lon
        FROM orders o
        JOIN users cu ON cu.id=o.customer_id
        LEFT JOIN couriers c ON c.id=o.courier_id
        LEFT JOIN users co ON co.id=c.user_id
        WHERE
            o.status!='closed'
            OR o.closed_at IS NULL
            OR o.closed_at>?
        ORDER BY o.id DESC
    """, (cutoff,)).fetchall()

    conn.close()

    return [dict(x) for x in rows]


@app.post("/api/admin/orders")
async def admin_create_order(
    data: OrderCreate,
    authorization: str = Header(default="")
):
    require_admin(authorization)
    phone = normalize_phone(data.phone)
    conn = db()
    customers = conn.execute("""
        SELECT id,name,phone,telegram_id FROM users
        WHERE role='customer' AND active=1
    """).fetchall()
    customer = next((row for row in customers if phones_equal(row["phone"], phone)), None)
    if not customer:
        conn.close()
        raise HTTPException(status_code=404, detail="Клиент с таким номером не найден")

    weight = float(data.weight_kg or 0)
    if weight < 0 or weight > MAX_ORDER_WEIGHT_KG:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Вес должен быть от 0 до {MAX_ORDER_WEIGHT_KG:g} кг")

    lat, lon = geocode_yerevan(data.address)
    restaurant_address = str(data.restaurant_address or "").strip()
    restaurant_lat, restaurant_lon = geocode_yerevan(restaurant_address) if restaurant_address else (None, None)
    bonus = calculate_courier_bonus(weight)

    cur = conn.execute("""
        INSERT INTO orders(
            customer_id,title,address,price,status,created_at,weight_kg,courier_bonus_amd,
            lat,lon,restaurant_address,restaurant_lat,restaurant_lon,restaurant_name,
            delivery_deadline,floor,entrance,apartment,intercom,recipient_name,
            delivery_note,payment_note,items_text,change_amount,route_no
        )
        VALUES(?,?,?,?,'new',?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        customer["id"], data.title.strip(), data.address.strip(), float(data.price),
        int(time.time()), weight, bonus, lat, lon, restaurant_address, restaurant_lat,
        restaurant_lon, data.restaurant_name.strip(), data.delivery_deadline.strip(),
        data.floor.strip(), data.entrance.strip(), data.apartment.strip(),
        data.intercom.strip(), data.recipient_name.strip(), data.delivery_note.strip(),
        data.payment_note.strip(), data.items_text.strip(), float(data.change_amount or 0),
        None
    ))
    conn.commit()
    order_id = cur.lastrowid
    conn.close()
    return {"ok": True, "id": order_id}


class RegisterData(BaseModel):
    name: str
    phone: str
    telegram_id: int | None = None

class AdminPersonalLogin(BaseModel):
    phone: str
    nickname: str

@app.post("/api/register")
async def register_customer(data: RegisterData):
    """Public customer registration: no admin approval and no PIN required."""
    name = str(data.name or "").strip()
    phone = normalize_phone(data.phone)
    if len(name) < 2:
        raise HTTPException(status_code=400, detail="Введите имя")
    if len(phone) < 5:
        raise HTTPException(status_code=400, detail="Введите корректный номер телефона")

    conn = db()
    try:
        existing = conn.execute("SELECT * FROM users WHERE phone=?", (phone,)).fetchone()
        if existing:
            if existing["active"] != 1:
                raise HTTPException(status_code=403, detail="Этот аккаунт деактивирован")
            if existing["role"] == "customer":
                token = create_session(existing["id"], "customer")
                if data.telegram_id:
                    conn.execute("UPDATE users SET telegram_id=? WHERE id=?", (int(data.telegram_id), existing["id"]))
                    conn.commit()
                return {"token": token, "role": "customer", "existing": True}
            raise HTTPException(status_code=409, detail="Этот номер уже используется другим типом аккаунта")

        cur = conn.execute("""
            INSERT INTO users(name,phone,pin_hash,pin_plain,role,telegram_id,created_at,active)
            VALUES(?,?,?,?,?,?,?,1)
        """, (
            name, phone, hash_pin(secrets.token_urlsafe(32)), "", "customer",
            int(data.telegram_id) if data.telegram_id else None, int(time.time())
        ))
        conn.commit()
        user_id = cur.lastrowid
        token = create_session(user_id, "customer")
        return {"token": token, "role": "customer", "existing": False}
    finally:
        conn.close()

@app.post("/api/admin/personal-login")
async def admin_personal_login(data: AdminPersonalLogin):
    phone = normalize_phone(data.phone)
    nickname = re.sub(r"\s+", " ", str(data.nickname or "").strip())
    expected = ADMIN_ACCOUNTS.get(phone)
    # +777 / +778 / +779 are intentional short internal admin logins,
    # so they must NOT pass through the normal 5+ digit phone validator.
    if not expected or nickname != expected:
        raise HTTPException(status_code=403, detail="Неверный номер или никнейм администратора")

    conn = db()
    try:
        admin = conn.execute("SELECT * FROM users WHERE phone=? AND role='admin'", (phone,)).fetchone()
        if not admin:
            cur = conn.execute("""
                INSERT INTO users(name,phone,pin_hash,pin_plain,role,created_at,active)
                VALUES(?,?,?,?,?,?,1)
            """, (nickname, phone, hash_pin(secrets.token_urlsafe(32)), "", "admin", int(time.time())))
            conn.commit()
            admin = conn.execute("SELECT * FROM users WHERE id=?", (cur.lastrowid,)).fetchone()
        else:
            conn.execute("UPDATE users SET name=?, active=1 WHERE id=?", (nickname, admin["id"]))
            conn.commit()
        token = create_session(admin["id"], "admin")
        return {"token": token, "role": "admin", "name": nickname}
    finally:
        conn.close()


@app.get("/api/customer/history")
async def customer_history(authorization: str = Header(default="")):
    user = require_user(authorization)
    if user["role"] != "customer":
        raise HTTPException(status_code=403, detail="Нет доступа")
    conn = db()
    rows = conn.execute("""
        SELECT o.*, c.id AS courier_record_id, cu.name AS courier_name
        FROM orders o
        LEFT JOIN couriers c ON c.id=o.courier_id
        LEFT JOIN users cu ON cu.id=c.user_id
        WHERE o.customer_id=?
        ORDER BY o.created_at DESC
    """, (user["user_id"],)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.get("/api/courier/history")
async def courier_history(authorization: str = Header(default="")):
    user = require_user(authorization)
    if user["role"] != "courier":
        raise HTTPException(status_code=403, detail="Нет доступа")
    conn = db()
    courier = conn.execute("SELECT id FROM couriers WHERE user_id=?", (user["user_id"],)).fetchone()
    if not courier:
        conn.close(); raise HTTPException(status_code=404, detail="Курьер не найден")
    rows = conn.execute("""
        SELECT o.*, u.name AS customer_name, u.phone AS customer_phone
        FROM orders o JOIN users u ON u.id=o.customer_id
        WHERE o.courier_id=? ORDER BY o.created_at DESC
    """, (courier["id"],)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.get("/api/courier/schedule")
async def courier_schedule(authorization: str = Header(default="")):
    user = require_user(authorization)
    if user["role"] != "courier":
        raise HTTPException(status_code=403, detail="Нет доступа")
    conn = db()
    courier = conn.execute("SELECT id FROM couriers WHERE user_id=? AND active=1", (user["user_id"],)).fetchone()
    if not courier:
        conn.close(); raise HTTPException(status_code=403, detail="Курьер не активен")
    rows = conn.execute("""
        SELECT id,title,address,delivery_deadline,status,route_no,created_at
        FROM orders WHERE courier_id=? AND status IN ('assigned','accepted','delivering')
        ORDER BY COALESCE(delivery_deadline,''), id
    """, (courier["id"],)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.get("/api/contact-admin")
async def contact_admin():
    return {"admins": [{"telegram_id": x} for x in ADMIN_IDS]}



def export_customers_xlsx():
    conn=db()
    rows=conn.execute("""
        SELECT u.id,u.name,u.phone,u.telegram_id,u.role,u.active,u.created_at,
               c.approved AS courier_approved,c.online AS courier_online
        FROM users u
        LEFT JOIN couriers c ON c.user_id=u.id
        WHERE u.role IN ('customer','courier')
        ORDER BY u.created_at ASC,u.id ASC
    """).fetchall()
    conn.close()
    wb=Workbook()
    ws=wb.active
    ws.title="Клиенты"
    ws.append(["ID","Имя","Телефон","Telegram ID","Роль","Активен","Курьер одобрен","Онлайн","Дата регистрации"])
    for cell in ws[1]:
        cell.font=Font(bold=True,color="FFFFFF")
        cell.fill=PatternFill("solid",fgColor="222222")
        cell.alignment=Alignment(horizontal="center",vertical="center")
    for r in rows:
        ws.append([
            r["id"],r["name"],r["phone"],r["telegram_id"] or "",r["role"],
            "Да" if r["active"] else "Нет",
            "Да" if r["courier_approved"] else ("" if r["role"]!="courier" else "Нет"),
            "Да" if r["courier_online"] else ("" if r["role"]!="courier" else "Нет"),
            datetime.fromtimestamp(r["created_at"]).strftime("%Y-%m-%d %H:%M")
        ])
    widths=[8,24,20,16,14,10,18,12,22]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[__import__("openpyxl").utils.get_column_letter(i)].width=w
    ws.freeze_panes="A2"
    ws.auto_filter.ref=ws.dimensions
    stamp=datetime.now().strftime("%Y%m%d_%H%M%S")
    path=EXPORT_DIR/f"sertal_customers_{stamp}.xlsx"
    wb.save(path)
    return path

def export_orders_xlsx():
    conn = db()
    rows = conn.execute("""
        SELECT o.*, cu.name AS courier_name, cu.phone AS courier_phone,
               u.name AS customer_name, u.phone AS customer_phone
        FROM orders o
        JOIN users u ON u.id=o.customer_id
        LEFT JOIN couriers c ON c.id=o.courier_id
        LEFT JOIN users cu ON cu.id=c.user_id
        ORDER BY o.created_at ASC, o.id ASC
    """).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Доставки"
    headers = [
        "№","Заказ","Ресторан","Курьер","Телефон курьера","Клиент","Телефон клиента",
        "Статус","Доставить до","Доставлен в","Адрес","Этаж","Подъезд","Кв/офис",
        "Домофон","Получатель","Комментарий","Оплата","Сдача","Состав заказа",
        "Сумма","Маршрут","Создан"
    ]
    ws.append(headers)
    thin = Side(style="thin", color="D9D9D9")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="222222")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)

    status_names = {
        "new":"Новый","assigned":"Назначен","accepted":"Принят",
        "delivering":"В доставке","delivered":"Доставлен","closed":"Закрыт"
    }
    for n, r in enumerate(rows, 1):
        created = datetime.fromtimestamp(r["created_at"]).strftime("%Y-%m-%d %H:%M")
        ws.append([
            n, r["id"], r["restaurant_name"] or r["restaurant_address"] or "",
            r["courier_name"] or "", r["courier_phone"] or "",
            r["customer_name"] or "", r["customer_phone"] or "",
            status_names.get(r["status"], r["status"]), r["delivery_deadline"] or "",
            r["delivered_at"] or "", r["address"] or "", r["floor"] or "",
            r["entrance"] or "", r["apartment"] or "", r["intercom"] or "",
            r["recipient_name"] or "", r["delivery_note"] or "", r["payment_note"] or "",
            r["change_amount"] or 0, r["items_text"] or r["title"] or "",
            r["price"] or 0, r["route_no"] or "", created
        ])

    widths = [6,9,20,20,17,20,17,14,16,17,34,9,10,12,14,18,30,18,10,36,12,10,18]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[__import__("openpyxl").utils.get_column_letter(idx)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.25,right=0.25,top=0.4,bottom=0.4,header=0.2,footer=0.2)
    ws.print_title_rows = "1:1"
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = EXPORT_DIR / f"orders_{stamp}.xlsx"
    wb.save(path)
    return path

@app.get("/api/admin/export-xlsx")
async def admin_export_xlsx(authorization: str = Header(default="")):
    require_admin(authorization)
    path = export_orders_xlsx()
    return FileResponse(path, filename=path.name,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.post("/api/admin/export-xlsx/telegram")
async def admin_export_xlsx_telegram(authorization: str = Header(default="")):
    require_admin(authorization)
    path = export_orders_xlsx()
    if not telegram_app:
        raise HTTPException(status_code=503, detail="Telegram-бот ещё не запущен")
    sent = 0
    for admin_id in ADMIN_IDS:
        try:
            with open(path, "rb") as fh:
                await telegram_app.bot.send_document(
                    chat_id=admin_id,
                    document=fh,
                    filename=path.name,
                    caption=f"📊 SERTAL DELIVERY · выгрузка заказов · {datetime.now().strftime('%d.%m.%Y %H:%M')}"
                )
            sent += 1
        except Exception as e:
            print("Telegram export:", admin_id, e)
    if not sent:
        raise HTTPException(status_code=502, detail="Не удалось отправить файл администраторам")
    return {"ok": True, "sent": sent, "filename": path.name}


async def export_loop():
    while True:
        try:
            export_orders_xlsx()
        except Exception as e:
            print("XLSX export:", e)
        await asyncio.sleep(12 * 60 * 60)

@app.post("/api/admin/orders/{order_id}/assign")
async def assign_order(
    order_id: int,
    data: AssignData,
    authorization: str = Header(default="")
):
    require_admin(authorization)

    conn = db()

    courier = conn.execute("""
        SELECT *
        FROM couriers
        WHERE id=?
        AND approved=1
        AND active=1
    """, (data.courier_id,)).fetchone()

    if not courier:
        conn.close()
        raise HTTPException(
            status_code=404,
            detail="Курьер не найден или не одобрен"
        )

    order = conn.execute(
        "SELECT * FROM orders WHERE id=? AND status NOT IN ('closed','delivered')",
        (order_id,)
    ).fetchone()

    if not order:
        conn.close()
        raise HTTPException(
            status_code=404,
            detail="Заказ не найден"
        )

    active_count = courier_active_order_count(conn, data.courier_id)
    if active_count >= BATCH_MAX_ORDERS:
        conn.close()
        raise HTTPException(status_code=400, detail="У курьера уже максимальные 2 активных заказа")

    if active_count == 1:
        distance = batch_distance_to_existing(conn, data.courier_id, order)
        if distance is not None and distance > BATCH_MAX_DISTANCE_KM:
            conn.close()
            raise HTTPException(
                status_code=400,
                detail=f"Второй заказ можно взять только если точки выполнения не дальше {BATCH_MAX_DISTANCE_KM:g} км (сейчас {distance:.2f} км)"
            )

    conn.execute("""
        UPDATE orders
        SET courier_id=?, status='assigned'
        WHERE id=?
        AND status!='closed'
    """, (
        data.courier_id,
        order_id
    ))

    conn.commit()
    conn.close()

    return {"ok": True}


@app.post("/api/admin/orders/{order_id}/close")
async def close_order(
    order_id: int,
    authorization: str = Header(default="")
):
    require_admin(authorization)

    now = int(time.time())

    conn = db()

    cur = conn.execute("""
        UPDATE orders
        SET status='closed',
            closed_at=?
        WHERE id=?
    """, (now, order_id))

    conn.commit()
    conn.close()

    if cur.rowcount == 0:
        raise HTTPException(
            status_code=404,
            detail="Заказ не найден"
        )

    return {
        "ok": True,
        "closed_at": now,
        "delete_after": now + 300
    }


# =========================================================
# CUSTOMER ORDERS
# =========================================================

@app.get("/api/customer/orders")
async def customer_orders(
    authorization: str = Header(default="")
):
    user = require_user(authorization)

    if user["role"] != "customer":
        raise HTTPException(
            status_code=403,
            detail="Нет доступа"
        )

    cleanup_old_closed()

    cutoff = order_cutoff()

    conn = db()

    rows = conn.execute("""
        SELECT
            o.*,
            c.lat,
            c.lon,
            co.name AS courier_name
        FROM orders o
        LEFT JOIN couriers c ON c.id=o.courier_id
        LEFT JOIN users co ON co.id=c.user_id
        WHERE o.customer_id=?
        AND (
            o.status!='closed'
            OR o.closed_at IS NULL
            OR o.closed_at>?
        )
        ORDER BY o.id DESC
    """, (
        user["user_id"],
        cutoff
    )).fetchall()

    conn.close()

    return [dict(x) for x in rows]


@app.post("/api/customer/orders/{order_id}/confirm")
async def customer_confirm(
    order_id: int,
    authorization: str = Header(default="")
):
    user = require_user(authorization)

    if user["role"] != "customer":
        raise HTTPException(
            status_code=403,
            detail="Нет доступа"
        )

    conn = db()

    order = conn.execute("""
        SELECT *
        FROM orders
        WHERE id=?
        AND customer_id=?
    """, (
        order_id,
        user["user_id"]
    )).fetchone()

    if not order:
        conn.close()
        raise HTTPException(
            status_code=404,
            detail="Заказ не найден"
        )

    if order["status"] != "delivered":
        conn.close()
        raise HTTPException(
            status_code=400,
            detail="Заказ ещё не доставлен"
        )

    now = int(time.time())

    conn.execute("""
        UPDATE orders
        SET customer_confirmed=1,
            status='closed',
            closed_at=?
        WHERE id=?
    """, (
        now,
        order_id
    ))

    conn.commit()
    conn.close()

    return {"ok": True}


# =========================================================
# RESTAURANTS / MARKETPLACE
# =========================================================

CATEGORY_SEARCHES = {
    "pizza": [
        "пицца доставка Ереван",
        "пицца рестораны Ереван заказать",
        "pizza delivery Yerevan Armenia",
        "pizza restaurants Yerevan Armenia",
    ],
    "sushi": [
        "суши доставка Ереван",
        "суши рестораны Ереван заказать",
        "sushi delivery Yerevan Armenia",
        "sushi restaurants Yerevan Armenia",
    ],
    "burgers": [
        "бургеры доставка Ереван",
        "бургер рестораны Ереван заказать",
        "burger delivery Yerevan Armenia",
    ],
    "shawarma": [
        "шаверма доставка Ереван",
        "шаурма рестораны Ереван заказать",
        "shawarma delivery Yerevan Armenia",
    ],
    "grocery": [
        "супермаркеты доставка Ереван",
        "продукты доставка Ереван",
        "grocery delivery Yerevan Armenia",
    ],
    "pharmacy": [
        "аптеки Ереван доставка",
        "pharmacy delivery Yerevan Armenia",
    ],
    "flowers": [
        "цветы доставка Ереван",
        "flower delivery Yerevan Armenia",
    ],
    "desserts": [
        "десерты доставка Ереван",
        "кондитерские Ереван доставка",
        "dessert delivery Yerevan Armenia",
    ],
    "coffee": [
        "кофе доставка Ереван",
        "кофейни Ереван доставка",
        "coffee delivery Yerevan Armenia",
    ],
}

CATEGORY_LABELS = {
    "pizza": "Пицца", "sushi": "Суши", "burgers": "Бургеры",
    "shawarma": "Шаурма", "grocery": "Супермаркеты",
    "pharmacy": "Аптеки", "flowers": "Цветы",
    "desserts": "Десерты", "coffee": "Кофе",
}


def _normalise_web_results(items):
    seen=set()
    out=[]
    for item in items or []:
        title=str(item.get("title") or "").strip()
        link=str(item.get("link") or "").strip()
        if not title or not link:
            continue
        key=re.sub(r"[^a-z0-9а-яё]+", "", (title+link).lower())
        if key in seen:
            continue
        seen.add(key)
        host=link.split("//",1)[-1].split("/",1)[0]
        out.append({
            "title": title[:180],
            "link": link[:700],
            "snippet": str(item.get("snippet") or "").strip()[:500],
            "displayLink": host[:160],
            "source": "internet",
        })
    return out


def gemini_google_search(query, limit=12, location="Yerevan, Armenia"):
    """Ask Gemini to search the live web through Google's built-in Search grounding.

    This is intentionally server-side: the browser never receives the Gemini key.
    If Gemini is unavailable or the key is absent, callers can fall back to the
    existing keyless DuckDuckGo discovery path.
    """
    key=GEMINI_API_KEY
    if not key:
        return {"ok": False, "text": "", "items": [], "queries": [], "error": "GEMINI_API_KEY not configured"}
    prompt=(
        "Ты поисковый агент SERTAL DELIVERY. Используй Google Search grounding и ищи в интернете прямо сейчас. "
        f"Локация: {location or 'Yerevan, Armenia'}. Запрос: {str(query or '').strip()[:500]}. "
        f"Найди до {max(3, min(int(limit or 12), 15))} реально существующих сервисов/заведений. "
        "Не выдумывай названия, адреса, цены или ссылки. Отдавай предпочтение официальным сайтам и актуальным страницам. "
        "Ответь кратко по-русски: название, что найдено, адрес/район если есть, и ссылка на источник."
    )
    body=json.dumps({
        "model": GEMINI_MODEL,
        "input": prompt,
        "tools": [{"type": "google_search"}],
    }).encode("utf-8")
    try:
        req=Request(
            "https://generativelanguage.googleapis.com/v1beta/interactions",
            data=body,
            headers={"Content-Type":"application/json", "x-goog-api-key":key},
            method="POST",
        )
        with urlopen(req, timeout=35) as response:
            payload=json.loads(response.read().decode("utf-8", "ignore"))
        text=""
        items=[]
        queries=[]
        for step in payload.get("steps", []) or []:
            st=step.get("type")
            if st=="google_search_call":
                queries.extend(step.get("arguments", {}).get("queries", []) or [])
            if st=="model_output":
                for block in step.get("content", []) or []:
                    if block.get("type")!="text":
                        continue
                    text += (str(block.get("text") or "").strip()+"\n")
                    for ann in block.get("annotations", []) or []:
                        if ann.get("type")!="url_citation":
                            continue
                        url=str(ann.get("url") or "").strip()
                        title=str(ann.get("title") or "").strip()
                        if not url or not url.startswith(("http://", "https://")):
                            continue
                        items.append({
                            "title": title[:180] or url.split("//",1)[-1].split("/",1)[0],
                            "link": url[:700],
                            "snippet": text.strip()[:500],
                            "displayLink": url.split("//",1)[-1].split("/",1)[0][:160],
                            "source": "gemini_google_search",
                        })
        items=_normalise_web_results(items)[:limit]
        return {"ok": True, "text": text.strip(), "items": items, "queries": queries[:12], "error": ""}
    except Exception as exc:
        print("GEMINI GOOGLE SEARCH:", exc)
        return {"ok": False, "text": "", "items": [], "queries": [], "error": str(exc)[:300]}


def category_internet_results(category, limit=30):
    queries=CATEGORY_SEARCHES.get(category, [f"{category} Ереван Armenia", f"{category} доставка Ереван"])
    merged=[]
    for query in queries:
        merged.extend(internet_search(query, max(8, min(12, limit))))
        if len(_normalise_web_results(merged)) >= limit:
            break
    return _normalise_web_results(merged)[:limit]


@app.get("/api/customer/search")
async def customer_search(q: str = "", authorization: str = Header(default="")):
    """Generic query search. No marketplace/service catalog is used here.

    Gemini gets Google's built-in Search grounding when GEMINI_API_KEY is set.
    Without Gemini, a keyless web-search fallback is used.
    """
    user = require_user(authorization)
    if user["role"] != "customer":
        raise HTTPException(403, "Нет доступа")
    query = str(q or "").strip()[:500]
    if not query:
        return {"ok": True, "answer": "", "sources": [], "grounded": False, "search_url": "https://www.google.com/search"}

    location = "Yerevan, Armenia"
    gemini = gemini_google_search(query, 12, location)
    if gemini["ok"]:
        return {
            "ok": True,
            "answer": gemini["text"],
            "sources": gemini["items"],
            "grounded": True,
            "provider": "Gemini + Google Search",
            "queries": gemini["queries"],
            "search_url": "https://www.google.com/search?q=" + quote(query),
            "message": "Gemini выполнил поиск по запросу через Google Search.",
        }

    items = _normalise_web_results(internet_search(query, 12))
    fallback_answer = "\n\n".join(
        f"{i+1}. {item.get('title','')}\n{item.get('snippet','')}"
        for i, item in enumerate(items[:8])
    )
    return {
        "ok": True,
        "answer": fallback_answer or "Поиск временно не вернул результатов.",
        "sources": items,
        "grounded": False,
        "provider": "Keyless internet fallback",
        "queries": [query],
        "search_url": "https://www.google.com/search?q=" + quote(query),
        "message": "Gemini не настроен или временно недоступен; показан резервный интернет-поиск.",
    }


@app.get("/api/customer/category-services")
async def customer_category_services(category: str = "", address: str = "", authorization: str = Header(default="")):
    user=require_user(authorization)
    if user["role"] != "customer":
        raise HTTPException(403, "Нет доступа")
    category=str(category or "").strip().lower()
    if category not in CATEGORY_SEARCHES:
        raise HTTPException(400, "Неизвестная категория")
    # First show SERTAL partners that can actually be ordered from inside the app.
    conn=db()
    restaurants=conn.execute("SELECT * FROM restaurants WHERE active=1 ORDER BY name").fetchall()
    partners=[]
    keywords={category, CATEGORY_LABELS.get(category, category)}
    for r in restaurants:
        items=conn.execute("SELECT id,name,description,price FROM menu_items WHERE restaurant_id=? AND active=1 ORDER BY id", (r["id"],)).fetchall()
        hay=(str(r["name"])+" "+str(r["address"])+" "+" ".join(str(i["name"]) for i in items)).lower()
        match=any(k.lower() in hay for k in keywords if k)
        # If the partner has a menu but no explicit category metadata, keep it for food categories
        # only when the category word is found. This avoids pretending every restaurant sells pizza/sushi.
        if match:
            partners.append({**dict(r), "items":[dict(i) for i in items], "source":"sertal"})
    conn.close()

    search_query=f"{CATEGORY_LABELS[category]} доставка рядом {address or 'Yerevan, Armenia'}"
    gemini=gemini_google_search(search_query, 20, address or "Yerevan, Armenia")
    web_items=gemini["items"] if gemini["ok"] else category_internet_results(category, 30)
    label=CATEGORY_LABELS[category]
    return {
        "ok": True,
        "category": category,
        "label": label,
        "city": "Yerevan",
        "address": str(address or "").strip()[:300],
        "partners": partners,
        "services": web_items,
        "total_found": len(partners)+len(web_items),
        "automatic": True,
        "ai_search": bool(gemini["ok"]),
        "ai_provider": "Gemini + Google Search" if gemini["ok"] else "Keyless internet fallback",
        "ai_summary": gemini["text"] if gemini["ok"] else "",
        "search_queries": gemini["queries"] or CATEGORY_SEARCHES[category],
        "notice": "Gemini автоматически ищет в Google Search и собирает актуальные внешние результаты. Наличие, цены и доставка внешних сервисов нужно проверять на их сайте." if gemini["ok"] else "Gemini не настроен, поэтому используется резервный интернет-поиск. Наличие и цены внешних сервисов нужно проверять на их сайте."
    }

@app.get("/api/version")
async def api_version():
    return {
        "name": "SERTAL DELIVERY",
        "release": "MEGA-2026.09",
        "marketplace": "disabled",
        "search": "generic-gemini-google-search",
        "google_required": False,
        "gemini_google_search": bool(GEMINI_API_KEY),
        "gemini_model": GEMINI_MODEL if GEMINI_API_KEY else "",
        "city": "Yerevan"
    }

@app.get("/api/public")
async def public_info():
    conn = db()
    rows = conn.execute("SELECT id,name,address,phone FROM restaurants WHERE active=1 ORDER BY name").fetchall()
    conn.close()
    return {"name":"SERTAL DELIVERY", "description":"Премиальная городская доставка еды, покупок и заказов.", "restaurants":[dict(r) for r in rows]}

@app.get("/api/customer/restaurants")
async def customer_restaurants(authorization: str = Header(default="")):
    user = require_user(authorization)
    if user["role"] != "customer":
        raise HTTPException(403, "Нет доступа")
    conn = db()
    restaurants = conn.execute("SELECT * FROM restaurants WHERE active=1 ORDER BY name").fetchall()
    result = []
    for r in restaurants:
        items = conn.execute("SELECT id,name,description,price FROM menu_items WHERE restaurant_id=? AND active=1 ORDER BY id", (r["id"],)).fetchall()
        result.append({**dict(r), "items":[dict(i) for i in items]})
    conn.close()
    return result

@app.post("/api/customer/orders/from-cart")
async def cart_order(data: CartOrderData, authorization: str = Header(default="")):
    user = require_user(authorization)
    if user["role"] != "customer":
        raise HTTPException(403, "Нет доступа")
    if not data.items or not data.address.strip():
        raise HTTPException(400, "Выберите товары и укажите адрес")
    payment = data.payment_method if data.payment_method in ("cash", "online") else "cash"
    conn = db()
    try:
        restaurant = conn.execute("SELECT * FROM restaurants WHERE id=? AND active=1", (data.restaurant_id,)).fetchone()
        if not restaurant:
            raise HTTPException(404, "Ресторан не найден")
        ids = [int(x.menu_item_id) for x in data.items]
        placeholders = ",".join("?" for _ in ids)
        rows = conn.execute(f"SELECT * FROM menu_items WHERE restaurant_id=? AND active=1 AND id IN ({placeholders})", [data.restaurant_id, *ids]).fetchall()
        by_id = {int(r["id"]): r for r in rows}
        parts=[]; total=0.0
        for item in data.items:
            r=by_id.get(int(item.menu_item_id))
            if not r or int(item.quantity) < 1:
                continue
            qty=min(int(item.quantity), 20)
            amount=float(r["price"])*qty
            total += amount
            parts.append(f"{r['name']} x{qty}")
        if not parts:
            raise HTTPException(400, "Выберите корректные товары")

        discount=0.0; promo_code=str(data.promo_code or "").strip().upper()
        if promo_code:
            promo=conn.execute("SELECT * FROM promos WHERE code=? AND active=1", (promo_code,)).fetchone()
            if not promo:
                raise HTTPException(400, "Промокод не найден или выключен")
            if float(total) < float(promo["min_order"] or 0):
                raise HTTPException(400, f"Минимальная сумма для промокода: {float(promo['min_order']):.0f} ֏")
            if promo["max_uses"] is not None and int(promo["used_count"] or 0) >= int(promo["max_uses"]):
                raise HTTPException(400, "Лимит использований промокода исчерпан")
            if promo["kind"] == "fixed":
                discount=min(total, float(promo["value"]))
            else:
                discount=min(total, total*float(promo["value"])/100.0)
            conn.execute("UPDATE promos SET used_count=used_count+1 WHERE id=?", (promo["id"],))
        final_total=max(0.0,total-discount)
        lat, lon = geocode_yerevan(data.address)
        cur=conn.execute("""
            INSERT INTO orders(customer_id,restaurant_id,restaurant_name,title,items_text,address,restaurant_address,price,payment_method,payment_status,status,created_at,lat,lon,delivery_note,promo_code,discount_amd)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (user["user_id"], restaurant["id"], restaurant["name"], "Заказ из ресторана", ", ".join(parts), data.address.strip(), restaurant["address"], final_total, payment, "pending" if payment == "online" else "cash", "new", int(time.time()), lat, lon, data.delivery_note.strip(), promo_code or None, discount))
        conn.commit()
        order_id=cur.lastrowid
        return {"ok":True,"order_id":order_id,"total":final_total,"subtotal":total,"discount":discount,"promo_code":promo_code}
    finally:
        conn.close()

@app.get("/api/customer/address")
async def customer_address(authorization: str = Header(default="")):
    user=require_user(authorization)
    conn=db(); row=conn.execute("SELECT default_address,default_lat,default_lon FROM users WHERE id=?",(user["user_id"],)).fetchone(); conn.close()
    return {"address":row["default_address"] if row else "","lat":row["default_lat"] if row else None,"lon":row["default_lon"] if row else None}

@app.post("/api/customer/address")
async def save_customer_address(data: CustomerAddressData, authorization: str = Header(default="")):
    user=require_user(authorization)
    address=str(data.address or "").strip()
    if len(address)<2 and (data.lat is None or data.lon is None):
        raise HTTPException(400,"Укажите адрес или разрешите геолокацию")
    conn=db(); conn.execute("UPDATE users SET default_address=?,default_lat=?,default_lon=? WHERE id=?",(address,data.lat,data.lon,user["user_id"])); conn.commit(); conn.close()
    return {"ok":True,"address":address,"lat":data.lat,"lon":data.lon}

@app.post("/api/customer/promo/check")
async def check_customer_promo(data: dict, authorization: str = Header(default="")):
    user=require_user(authorization)
    if user["role"]!="customer": raise HTTPException(403,"Нет доступа")
    code=str(data.get("code") or "").strip().upper(); total=float(data.get("total") or 0)
    if not code: raise HTTPException(400,"Введите промокод")
    conn=db(); promo=conn.execute("SELECT * FROM promos WHERE code=? AND active=1",(code,)).fetchone(); conn.close()
    if not promo: raise HTTPException(404,"Промокод не найден")
    if total<float(promo["min_order"] or 0): raise HTTPException(400,f"Минимальная сумма: {float(promo['min_order']):.0f} ֏")
    if promo["max_uses"] is not None and int(promo["used_count"] or 0)>=int(promo["max_uses"]): raise HTTPException(400,"Лимит использований исчерпан")
    discount=min(total, float(promo["value"]) if promo["kind"]=="fixed" else total*float(promo["value"])/100.0)
    return {"ok":True,"code":code,"discount":discount,"total":max(0,total-discount),"kind":promo["kind"],"value":promo["value"]}

@app.get("/api/admin/restaurants")
async def admin_restaurants(authorization: str = Header(default="")):
    require_admin(authorization)
    conn = db(); restaurants = conn.execute("SELECT * FROM restaurants ORDER BY id DESC").fetchall()
    result = []
    for r in restaurants:
        items = conn.execute("SELECT * FROM menu_items WHERE restaurant_id=? ORDER BY id DESC", (r["id"],)).fetchall()
        result.append({**dict(r), "items":[dict(i) for i in items]})
    conn.close(); return result

@app.post("/api/admin/restaurants")
async def add_restaurant(data: RestaurantCreate, authorization: str = Header(default="")):
    require_admin(authorization)
    name = data.name.strip()
    if not name: raise HTTPException(400, "Введите название ресторана")
    conn = db(); cur = conn.execute("INSERT INTO restaurants(name,address,phone,active,created_at) VALUES(?,?,?,?,?)", (name, data.address.strip(), data.phone.strip(), 1, int(time.time()))); conn.commit(); rid = cur.lastrowid; conn.close()
    return {"ok":True,"id":rid}

@app.post("/api/admin/restaurants/menu")
async def add_menu(data: MenuItemCreate, authorization: str = Header(default="")):
    require_admin(authorization)
    if data.price < 0: raise HTTPException(400, "Цена не может быть отрицательной")
    conn = db(); exists = conn.execute("SELECT id FROM restaurants WHERE id=?", (data.restaurant_id,)).fetchone()
    if not exists: conn.close(); raise HTTPException(404, "Ресторан не найден")
    cur = conn.execute("INSERT INTO menu_items(restaurant_id,name,description,price,active,created_at) VALUES(?,?,?,?,1,?)", (data.restaurant_id, data.name.strip(), data.description.strip(), data.price, int(time.time()))); conn.commit(); mid = cur.lastrowid; conn.close()
    return {"ok":True,"id":mid}

@app.post("/api/admin/restaurants/{restaurant_id}/toggle")
async def toggle_restaurant(restaurant_id: int, authorization: str = Header(default="")):
    require_admin(authorization)
    conn = db(); conn.execute("UPDATE restaurants SET active=CASE active WHEN 1 THEN 0 ELSE 1 END WHERE id=?", (restaurant_id,)); conn.commit(); conn.close(); return {"ok":True}

@app.post("/api/admin/customers/{customer_id}/promote-courier")
async def promote_courier(customer_id: int, authorization: str = Header(default="")):
    require_admin(authorization)
    conn = db(); user = conn.execute("SELECT * FROM users WHERE id=? AND role='customer' AND active=1", (customer_id,)).fetchone()
    if not user: conn.close(); raise HTTPException(404, "Покупатель не найден")
    conn.execute("UPDATE users SET role='courier' WHERE id=?", (customer_id,))
    existing = conn.execute("SELECT id FROM couriers WHERE user_id=?", (customer_id,)).fetchone()
    if existing:
        conn.execute("UPDATE couriers SET approved=1,active=1 WHERE user_id=?", (customer_id,))
    else:
        conn.execute("INSERT INTO couriers(user_id,approved,active,online) VALUES(?,1,1,0)", (customer_id,))
    conn.execute("DELETE FROM sessions WHERE user_id=?", (customer_id,)); conn.commit(); conn.close(); log_action(customer_id, "promote_courier", ""); return {"ok":True}

@app.get("/api/admin/map-orders")
async def admin_map_orders(authorization: str = Header(default="")):
    require_admin(authorization)
    conn = db()
    orders = conn.execute("SELECT id,title,address,lat,lon,status,price FROM orders WHERE lat IS NOT NULL AND lon IS NOT NULL ORDER BY id DESC LIMIT 500").fetchall()
    couriers = conn.execute("SELECT c.id,u.name,c.lat,c.lon,c.online FROM couriers c JOIN users u ON u.id=c.user_id WHERE c.active=1 AND c.lat IS NOT NULL AND c.lon IS NOT NULL").fetchall()
    conn.close()
    return {"orders":[dict(x) for x in orders], "couriers":[dict(x) for x in couriers]}

# =========================================================
# COURIER
# =========================================================

@app.get("/api/courier/orders")
async def courier_orders(
    authorization: str = Header(default="")
):
    user = require_user(authorization)

    if user["role"] != "courier":
        raise HTTPException(
            status_code=403,
            detail="Нет доступа"
        )

    cleanup_old_closed()

    cutoff = order_cutoff()

    conn = db()

    courier = conn.execute("""
        SELECT *
        FROM couriers
        WHERE user_id=?
    """, (user["user_id"],)).fetchone()

    if not courier or not courier["active"]:
        conn.close()
        raise HTTPException(
            status_code=403,
            detail="Курьер деактивирован"
        )

    rows = conn.execute("""
        SELECT
            o.*,
            u.name AS customer_name,
            u.phone AS customer_phone
        FROM orders o
        JOIN users u ON u.id=o.customer_id
        WHERE o.courier_id=?
        AND (
            o.status!='closed'
            OR o.closed_at IS NULL
            OR o.closed_at>?
        )
        ORDER BY o.id DESC
    """, (
        courier["id"],
        cutoff
    )).fetchall()

    conn.close()

    return {
        "online": bool(courier["online"]),
        "orders": [dict(x) for x in rows]
    }


@app.get("/api/courier/batch-options")
async def courier_batch_options(
    authorization: str = Header(default="")
):
    user = require_user(authorization)
    if user["role"] != "courier":
        raise HTTPException(status_code=403, detail="Нет доступа")

    conn = db()
    courier = conn.execute(
        "SELECT * FROM couriers WHERE user_id=? AND active=1",
        (user["user_id"],)
    ).fetchone()
    if not courier:
        conn.close()
        raise HTTPException(status_code=403, detail="Курьер не активен")

    active_count = courier_active_order_count(conn, courier["id"])
    rows = conn.execute("""
        SELECT o.*, u.name AS customer_name, u.phone AS customer_phone
        FROM orders o
        JOIN users u ON u.id=o.customer_id
        WHERE o.status='assigned' AND o.courier_id=?
        ORDER BY o.id DESC
    """, (courier["id"],)).fetchall()

    options = []
    for row in rows:
        d = batch_distance_to_existing(conn, courier["id"], row)
        options.append({
            **dict(row),
            "batch_allowed": active_count < BATCH_MAX_ORDERS and (
                active_count == 0 or d is None or d <= BATCH_MAX_DISTANCE_KM
            ),
            "distance_km": d
        })
    conn.close()
    return {
        "active_orders": active_count,
        "max_orders": BATCH_MAX_ORDERS,
        "max_distance_km": BATCH_MAX_DISTANCE_KM,
        "options": options
    }


@app.post("/api/courier/online")
async def courier_online(
    data: OnlineData,
    authorization: str = Header(default="")
):
    user = require_user(authorization)

    if user["role"] != "courier":
        raise HTTPException(
            status_code=403,
            detail="Нет доступа"
        )

    conn = db()

    courier = conn.execute(
        "SELECT * FROM couriers WHERE user_id=?",
        (user["user_id"],)
    ).fetchone()

    if not courier or not courier["active"]:
        conn.close()
        raise HTTPException(
            status_code=403,
            detail="Курьер деактивирован"
        )

    conn.execute("""
        UPDATE couriers
        SET online=?
        WHERE id=?
    """, (
        1 if data.online else 0,
        courier["id"]
    ))

    conn.commit()
    conn.close()

    return {"ok": True}


@app.post("/api/courier/orders/{order_id}/accept")
async def courier_accept(
    order_id: int,
    authorization: str = Header(default="")
):
    user = require_user(authorization)

    if user["role"] != "courier":
        raise HTTPException(
            status_code=403,
            detail="Нет доступа"
        )

    conn = db()

    courier = conn.execute(
        "SELECT * FROM couriers WHERE user_id=? AND active=1",
        (user["user_id"],)
    ).fetchone()

    if not courier:
        conn.close()
        raise HTTPException(
            status_code=403,
            detail="Курьер не активен"
        )

    order = conn.execute(
        "SELECT * FROM orders WHERE id=? AND courier_id=?",
        (order_id, courier["id"])
    ).fetchone()

    if not order or order["status"] != "assigned":
        conn.close()
        raise HTTPException(status_code=400, detail="Заказ нельзя принять")

    active_count = courier_active_order_count(conn, courier["id"])
    if active_count >= BATCH_MAX_ORDERS:
        conn.close()
        raise HTTPException(status_code=400, detail="Можно одновременно иметь максимум 2 активных заказа")

    if active_count == 1:
        distance = batch_distance_to_existing(conn, courier["id"], order)
        if distance is not None and distance > BATCH_MAX_DISTANCE_KM:
            conn.close()
            raise HTTPException(
                status_code=400,
                detail=f"Второй заказ нельзя принять: точки дальше 1 км ({distance:.2f} км)"
            )

    cur = conn.execute("""
        UPDATE orders
        SET status='accepted'
        WHERE id=?
        AND courier_id=?
        AND status='assigned'
    """, (
        order_id,
        courier["id"]
    ))

    conn.commit()
    conn.close()

    if cur.rowcount == 0:
        raise HTTPException(
            status_code=400,
            detail="Заказ нельзя принять"
        )

    return {"ok": True}


@app.post("/api/courier/orders/{order_id}/start")
async def courier_start(
    order_id: int,
    authorization: str = Header(default="")
):
    user = require_user(authorization)

    if user["role"] != "courier":
        raise HTTPException(
            status_code=403,
            detail="Нет доступа"
        )

    conn = db()

    courier = conn.execute(
        "SELECT * FROM couriers WHERE user_id=? AND active=1",
        (user["user_id"],)
    ).fetchone()

    if not courier:
        conn.close()
        raise HTTPException(
            status_code=403,
            detail="Курьер не активен"
        )

    cur = conn.execute("""
        UPDATE orders
        SET status='delivering'
        WHERE id=?
        AND courier_id=?
        AND status='accepted'
    """, (
        order_id,
        courier["id"]
    ))

    conn.commit()
    conn.close()

    if cur.rowcount == 0:
        raise HTTPException(
            status_code=400,
            detail="Нельзя начать эту доставку"
        )

    return {"ok": True}


@app.post("/api/courier/orders/{order_id}/complete")
async def courier_complete(
    order_id: int,
    authorization: str = Header(default="")
):
    user = require_user(authorization)

    if user["role"] != "courier":
        raise HTTPException(
            status_code=403,
            detail="Нет доступа"
        )

    conn = db()

    courier = conn.execute(
        "SELECT * FROM couriers WHERE user_id=? AND active=1",
        (user["user_id"],)
    ).fetchone()

    if not courier:
        conn.close()
        raise HTTPException(
            status_code=403,
            detail="Курьер не активен"
        )

    cur = conn.execute("""
        UPDATE orders
        SET status='delivered', delivered_at=?
        WHERE id=?
        AND courier_id=?
        AND status='delivering'
    """, (
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        order_id,
        courier["id"]
    ))

    conn.commit()
    conn.close()

    if cur.rowcount == 0:
        raise HTTPException(
            status_code=400,
            detail="Нельзя завершить эту доставку"
        )

    return {"ok": True}


@app.post("/api/courier/location")
async def courier_location(
    data: LocationData,
    authorization: str = Header(default="")
):
    user = require_user(authorization)

    if user["role"] != "courier":
        raise HTTPException(
            status_code=403,
            detail="Нет доступа"
        )

    conn = db()

    courier = conn.execute(
        "SELECT * FROM couriers WHERE user_id=? AND active=1",
        (user["user_id"],)
    ).fetchone()

    if not courier:
        conn.close()
        raise HTTPException(
            status_code=403,
            detail="Курьер не активен"
        )

    conn.execute("""
        UPDATE couriers
        SET lat=?, lon=?, updated_at=?
        WHERE id=?
    """, (
        data.lat,
        data.lon,
        int(time.time()),
        courier["id"]
    ))

    conn.commit()
    conn.close()

    return {"ok": True}


# =========================================================
# SUPPORT CHAT
# =========================================================

def add_chat_message(user_id, sender_role, text="", file_name=None):
    conn = db()
    cur = conn.execute("""
        INSERT INTO chat_messages(user_id,sender_role,text,file_name,created_at)
        VALUES(?,?,?,?,?)
    """, (user_id, sender_role, text or "", file_name, int(time.time())))
    conn.commit()
    message_id = cur.lastrowid
    conn.close()
    return message_id


@app.get("/api/chat/messages")
async def get_chat_messages(authorization: str = Header(default="")):
    user = require_user(authorization)
    conn = db()
    rows = conn.execute("""
        SELECT id,user_id,sender_role,text,file_name,created_at
        FROM chat_messages
        WHERE user_id=?
        ORDER BY id ASC
    """, (user["user_id"],)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


async def notify_admins_about_user_message(user, text, file_path=None, file_name=None):
    if not telegram_app:
        return
    body = (
        f"💬 SERTAL DELIVERY\n"
        f"👤 {user['name']}\n"
        f"📱 {user['phone']}\n"
        f"🆔 user_id: {user['user_id']}\n\n"
        f"{text or '📎 Файл'}"
    )

    # Preferred mode: one dedicated Telegram support group.
    # Admins answer by replying to the bot's message in that group.
    support_group_id = get_support_group_chat_id()
    targets = [support_group_id] if support_group_id else list(ADMIN_IDS)
    for target_chat_id in targets:
        if not target_chat_id:
            continue
        try:
            if file_path:
                with open(file_path, "rb") as fh:
                    sent = await telegram_app.bot.send_document(
                        chat_id=target_chat_id,
                        document=fh,
                        filename=file_name or Path(file_path).name,
                        caption=body
                    )
            else:
                sent = await telegram_app.bot.send_message(chat_id=target_chat_id, text=body)

            conn = db()
            conn.execute("""
                INSERT OR IGNORE INTO admin_message_bridge(
                    admin_telegram_id,telegram_message_id,user_id,created_at
                ) VALUES(?,?,?,?)
            """, (int(target_chat_id), sent.message_id, user["user_id"], int(time.time())))
            conn.commit()
            conn.close()
        except Exception as e:
            print("Notify support target:", target_chat_id, e)


@app.post("/api/chat/send")
async def send_chat_message(data: ChatMessageData, authorization: str = Header(default="")):
    user = require_user(authorization)
    text = (data.text or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="Введите сообщение")
    msg_id = add_chat_message(user["user_id"], user["role"], text)
    await notify_admins_about_user_message(user, text)
    return {"ok": True, "id": msg_id}


@app.post("/api/chat/upload")
async def upload_chat_file(file: UploadFile = File(...), authorization: str = Header(default="")):
    user = require_user(authorization)
    original_name = Path(file.filename or "file").name
    safe_name = re.sub(r"[^A-Za-z0-9А-Яа-я._-]+", "_", original_name)[:180]
    path = UPLOAD_DIR / f"{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}_{safe_name}"
    with open(path, "wb") as out:
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            out.write(chunk)

    msg_id = add_chat_message(user["user_id"], user["role"], "📎 Файл", original_name)
    await notify_admins_about_user_message(user, "📎 Файл из веб-чата", str(path), original_name)
    return {"ok": True, "id": msg_id, "file_name": original_name}



# =========================================================
# TELEGRAM VISUAL REPORTS / AI
# =========================================================

def _font(size, bold=False):
    candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/dejavu/DejaVuSans.ttf",
    ]
    for path in candidates:
        if Path(path).exists():
            return ImageFont.truetype(path, size)
    return ImageFont.load_default()


def _make_report_image(title, columns, rows, path, subtitle="SERTAL DELIVERY"):
    """Render a compact, phone-friendly table as PNG. This is deterministic, not fake AI."""
    columns = [str(x) for x in columns]
    rows = [[str(x if x is not None else "") for x in row] for row in rows]
    scale = 1
    width = 1400
    margin = 40
    header_h = 105
    row_h = 58
    max_rows = 18
    rows = rows[:max_rows]
    col_widths = [max(130, min(360, 120 + max([len(r[i]) if i < len(r) else 0 for r in rows] + [len(columns[i])]) * 9)) for i in range(len(columns))]
    total = sum(col_widths) + margin * 2
    width = max(width, min(total, 2200))
    # Scale columns to fit the canvas.
    if total > width:
        factor = (width - margin*2) / sum(col_widths)
        col_widths = [max(100, int(x*factor)) for x in col_widths]
        total = sum(col_widths) + margin*2
        width = total
    height = header_h + (len(rows)+1)*row_h + margin
    img = Image.new("RGB", (width, height), (9, 9, 9))
    draw = ImageDraw.Draw(img)
    gold=(212,175,55); white=(245,241,230); muted=(150,142,125); panel=(20,19,16); line=(67,57,31)
    draw.rounded_rectangle((18,18,width-18,height-18), radius=28, fill=panel, outline=gold, width=2)
    draw.text((45,35), title, font=_font(32, True), fill=white)
    draw.text((45,75), subtitle, font=_font(18), fill=muted)
    y=header_h
    x=margin
    for i,col in enumerate(columns):
        cw=col_widths[i]
        draw.rectangle((x,y,x+cw,y+row_h), fill=(30,27,20), outline=line)
        draw.text((x+12,y+17), col[:28], font=_font(18, True), fill=gold)
        x += cw
    y += row_h
    for ri,row in enumerate(rows):
        x=margin
        for i in range(len(columns)):
            cw=col_widths[i]
            val=row[i] if i < len(row) else ""
            draw.rectangle((x,y,x+cw,y+row_h), fill=(14,14,13) if ri%2==0 else (18,18,17), outline=line)
            # one-line truncation for mobile readability
            max_chars=max(8,int(cw/10))
            draw.text((x+12,y+18), val[:max_chars], font=_font(17), fill=white)
            x += cw
        y += row_h
    img.save(path, "PNG", optimize=True)
    return path


def make_customers_report_image():
    conn=db()
    rows=conn.execute("""
        SELECT u.id,u.name,u.phone,u.role,u.active,
               COALESCE(c.approved,0) AS approved,
               COALESCE(c.online,0) AS online
        FROM users u LEFT JOIN couriers c ON c.user_id=u.id
        WHERE u.role IN ('customer','courier')
        ORDER BY u.id DESC LIMIT 40
    """).fetchall()
    conn.close()
    stamp=datetime.now().strftime("%Y%m%d_%H%M%S")
    path=EXPORT_DIR/f"sertal_clients_{stamp}.png"
    data=[]
    for r in rows:
        data.append([r['id'],r['name'],r['phone'],r['role'],"Да" if r['active'] else "Нет", "Да" if r['approved'] else "Нет", "Да" if r['online'] else "Нет"])
    return _make_report_image("Клиентская база", ["ID","Имя","Телефон","Роль","Активен","Одобрен","Онлайн"], data, path, f"SERTAL DELIVERY · {datetime.now().strftime('%d.%m.%Y %H:%M')}")


def make_orders_report_image():
    conn=db()
    rows=conn.execute("""
        SELECT o.id,u.name AS customer_name,o.title,o.address,o.status,o.price
        FROM orders o JOIN users u ON u.id=o.customer_id
        ORDER BY o.id DESC LIMIT 25
    """).fetchall()
    conn.close()
    stamp=datetime.now().strftime("%Y%m%d_%H%M%S")
    path=EXPORT_DIR/f"sertal_orders_{stamp}.png"
    data=[]
    for r in rows:
        data.append([r['id'],r['customer_name'],r['title'],r['address'],r['status'],f"{float(r['price'] or 0):.0f} ֏"])
    return _make_report_image("Заказы", ["№","Клиент","Что доставить","Адрес","Статус","Сумма"], data, path, f"SERTAL DELIVERY · {datetime.now().strftime('%d.%m.%Y %H:%M')}")


def make_random_receipt_image():
    import random
    names=["Кофе","Бургер","Пицца","Салат","Десерт","Напиток"]
    items=[]
    total=0
    for _ in range(random.randint(2,5)):
        name=random.choice(names); qty=random.randint(1,2); price=random.choice([900,1200,1500,1800,2200,2600]); amount=qty*price; total+=amount
        items.append([name,qty,f"{amount:,}".replace(',',' ')+" ֏"])
    stamp=datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    path=EXPORT_DIR/f"sertal_random_receipt_{stamp}.png"
    img=Image.new("RGB",(900,1050),(8,8,8)); d=ImageDraw.Draw(img)
    d.rounded_rectangle((35,35,865,1015),radius=35,fill=(20,18,14),outline=(212,175,55),width=3)
    d.text((70,70),"SERTAL DELIVERY",font=_font(34,True),fill=(212,175,55))
    d.text((70,120),"Демонстрационный чек",font=_font(20),fill=(170,160,140))
    y=185
    for name,qty,amount in items:
        d.text((70,y),f"{name} × {qty}",font=_font(25),fill=(245,241,230))
        d.text((650,y),amount,font=_font(25,True),fill=(245,241,230))
        y+=62
    d.line((70,y+10,830,y+10),fill=(67,57,31),width=2); y+=48
    d.text((70,y),"ИТОГО",font=_font(28,True),fill=(212,175,55))
    d.text((650,y),f"{total:,}".replace(',',' ')+" ֏",font=_font(28,True),fill=(245,241,230))
    y+=85
    d.text((70,y),f"Заказ №{random.randint(1000,9999)}",font=_font(20),fill=(150,142,125))
    d.text((70,y+35),datetime.now().strftime("%d.%m.%Y %H:%M"),font=_font(20),fill=(150,142,125))
    img.save(path,"PNG",optimize=True)
    return path


def internet_search(query, limit=6):
    """Keyless web discovery used by the marketplace and AI.

    The app intentionally does not require GOOGLE_API_KEY/GOOGLE_CX for the basic
    discovery flow. DuckDuckGo HTML is used server-side so the browser never
    needs to bypass CORS or expose a search key. Results are always treated as
    external leads, not as SERTAL partners.
    """
    from html.parser import HTMLParser

    class ResultParser(HTMLParser):
        def __init__(self):
            super().__init__()
            self.results=[]
            self.current=None
            self.capture=None

        def handle_starttag(self, tag, attrs):
            attrs=dict(attrs)
            classes=set(str(attrs.get("class", "")).split())
            if tag=="a" and "result__a" in classes:
                self.current={"title":"", "link":attrs.get("href", ""), "snippet":""}
                self.capture="title"
                return
            if self.current and tag in ("a","div") and "result__snippet" in classes:
                self.capture="snippet"

        def handle_data(self, data):
            if not self.current or not self.capture:
                return
            value=" ".join(str(data).split())
            if not value:
                return
            self.current[self.capture]=(self.current.get(self.capture, "")+" "+value).strip()

        def handle_endtag(self, tag):
            if not self.current:
                return
            if tag=="a" and self.capture=="title":
                self.capture=None
            # DDG wraps each result in a result container. Also flush defensively
            # when enough information is present so small HTML changes do not
            # make the whole search empty.
            if self.current.get("title") and self.current.get("link") and len(self.results)<limit:
                if tag in ("div", "article") and self.capture is None:
                    self.results.append(self.current)
                    self.current=None

    query=str(query or "").strip()[:220]
    if not query:
        return []
    try:
        url="https://html.duckduckgo.com/html/?q="+quote(query)
        req=Request(url,headers={"User-Agent":"Mozilla/5.0 SERTAL-DELIVERY/2026"})
        with urlopen(req,timeout=10) as response:
            html=response.read().decode("utf-8","ignore")
        parser=ResultParser(); parser.feed(html)
        # A second lightweight regex fallback handles DDG markup variants.
        if not parser.results:
            anchors=re.findall(r'<a[^>]+class=["\'][^"\']*result__a[^"\']*["\'][^>]*href=["\']([^"\']+)["\'][^>]*>(.*?)</a>', html, re.I|re.S)
            for link,title_html in anchors[:limit]:
                title=re.sub(r"<[^>]+>"," ",title_html)
                title=" ".join(re.sub(r"&[^;]+;", " ", title).split())
                parser.results.append({"title":title,"link":link,"snippet":""})
        return parser.results[:limit]
    except Exception as exc:
        print("WEB SEARCH:",exc)
        return []


async def ai_group_answer(question):
    """Shared AI answer for admin Web App and Telegram group."""
    question=str(question or "").strip()
    if not question:
        return "Укажите вопрос после !ai. Например: !ai сколько активных заказов?"
    conn=db()
    stats=conn.execute("""
        SELECT COUNT(*) total,
        SUM(CASE WHEN status='new' THEN 1 ELSE 0 END) new_count,
        SUM(CASE WHEN status IN ('assigned','accepted','delivering') THEN 1 ELSE 0 END) active_count,
        SUM(CASE WHEN status='delivered' THEN 1 ELSE 0 END) delivered_count,
        COALESCE(SUM(price),0) revenue
        FROM orders WHERE status!='closed'
    """).fetchone()
    customers=conn.execute("SELECT COUNT(*) n FROM users WHERE role IN ('customer','courier') AND active=1").fetchone()["n"]
    couriers=conn.execute("SELECT COUNT(*) n FROM couriers WHERE active=1 AND approved=1").fetchone()["n"]
    conn.close()
    context=(f"Активные данные SERTAL DELIVERY: заказов={stats['total'] or 0}, новых={stats['new_count'] or 0}, в работе={stats['active_count'] or 0}, доставлено={stats['delivered_count'] or 0}, сумма активных заказов={float(stats['revenue'] or 0):.0f} AMD, активных участников={customers}, одобренных курьеров={couriers}.")
    web_results=internet_search(question,6)
    web_context=""
    if web_results:
        web_context="\nИнтернет-результаты (проверяй по ссылкам, не выдумывай):\n"+"\n".join(f"- {x['title']} | {x['link']} | {x['snippet']}" for x in web_results)
    key=os.getenv("OPENAI_API_KEY","").strip()
    model=os.getenv("OPENAI_MODEL","gpt-5.6-luna")
    if key:
        try:
            prompt=("Ты операционный ИИ-помощник SERTAL DELIVERY. Отвечай по-русски, коротко и конкретно. "
                    "Не выдумывай факты. Если вопрос про данные компании, используй только переданный контекст. "
                    "Не выполняй опасные или необратимые действия через текст. " + context + " Вопрос: " + question)
            body=json.dumps({"model":model,"input":prompt}).encode("utf-8")
            req=Request("https://api.openai.com/v1/responses",data=body,headers={"Content-Type":"application/json","Authorization":f"Bearer {key}"},method="POST")
            with urlopen(req,timeout=25) as response:
                result=json.loads(response.read().decode("utf-8"))
            answer=result.get("output_text") or ""
            if not answer:
                parts=[]
                for item in result.get("output",[]):
                    for content in item.get("content",[]):
                        if content.get("type")=="output_text": parts.append(content.get("text", ""))
                answer="\n".join(parts).strip()
            if answer: return answer
        except Exception as exc:
            print("GROUP AI:",exc)
    q=question.lower()
    if any(x in q for x in ("сколько","заказ","заказы","статист")):
        return (f"Заказов: {stats['total'] or 0}\nНовых: {stats['new_count'] or 0}\nВ работе: {stats['active_count'] or 0}\nДоставлено: {stats['delivered_count'] or 0}\nСумма активных: {float(stats['revenue'] or 0):.0f} ֏")
    if any(x in q for x in ("клиент","покупател")):
        return f"Активных клиентов/курьеров в базе: {customers}."
    if "курьер" in q:
        return f"Одобренных активных курьеров: {couriers}."
    if web_results:
        return "Нашёл в интернете:\n\n"+"\n\n".join(f"• {x['title']}\n{x['snippet']}\n{x['link']}" for x in web_results)
    return "ИИ подключён, но интернет-поиск сейчас недоступен. Попробуйте ещё раз."

# =========================================================
# TELEGRAM BOT
# =========================================================

telegram_app = None


async def start_command(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    keyboard = [
        [
            InlineKeyboardButton(
                "🍽 Открыть приложение",
                web_app=WebAppInfo(
                    url=WEB_APP_URL
                )
            )
        ]
    ]

    await update.message.reply_text(
        "🍽 SERTAL DELIVERY\n\n"
        "Нажмите кнопку ниже, чтобы открыть приложение.",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def random_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "📱 Вход в SERTAL DELIVERY выполняется по номеру телефона, который добавил администратор."
    )


async def contact_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    contact=update.message.contact
    phone=normalize_phone(contact.phone_number)
    telegram_id=update.effective_user.id
    conn=db()
    row=conn.execute("SELECT id,name,role,active FROM users WHERE phone=?",(phone,)).fetchone()
    if row and row["active"]:
        conn.execute("UPDATE users SET telegram_id=? WHERE id=?",(telegram_id,row["id"]))
        conn.commit()
    conn.close()
    if not row or not row["active"]:
        await update.message.reply_text("Клиент с таким номером не найден в SERTAL DELIVERY.")
        return
    await update.message.reply_text("Номер подтверждён. Откройте SERTAL DELIVERY и войдите по номеру телефона.")


async def logsfile_direct(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message or not update.effective_user or update.effective_user.id not in ADMIN_IDS:
        return
    path = export_customers_xlsx()
    try:
        with open(path, "rb") as fh:
            await update.message.reply_document(document=fh, filename=path.name, caption="SERTAL DELIVERY — клиентская база")
    except Exception as e:
        await update.message.reply_text(f"Ошибка выгрузки: {e}")


async def promo_group_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    message=update.message
    if not message or not update.effective_user or update.effective_user.id not in ADMIN_IDS:
        return
    if message.chat.type == "private":
        await message.reply_text("Команды промокодов предназначены для рабочей Telegram-группы.")
        return
    text=(message.text or "").strip()
    parts=text.split()
    low=parts[0].lower() if parts else ""
    conn=db()
    try:
        if low in ("!promos","!promo_list","!промокоды"):
            rows=conn.execute("SELECT code,kind,value,min_order,max_uses,used_count,active FROM promos ORDER BY id DESC").fetchall()
            if not rows:
                await message.reply_text("Промокодов пока нет.\nПример: !promo SERTAL10 10%")
                return
            lines=[f"{r['code']} — {'скидка '+str(r['value'])+'%' if r['kind']=='percent' else '−'+str(r['value'])+' ֏'} — мин. {float(r['min_order'] or 0):.0f} ֏ — {'ON' if r['active'] else 'OFF'} — {r['used_count'] or 0}/{r['max_uses'] if r['max_uses'] is not None else '∞'}" for r in rows]
            await message.reply_text("🎟 ПРОМОКОДЫ\n\n"+"\n".join(lines))
            return
        if low in ("!delpromo","!deletepromo","!удалитьпромо"):
            if len(parts)<2: await message.reply_text("Формат: !delpromo CODE"); return
            code=parts[1].upper(); cur=conn.execute("UPDATE promos SET active=0 WHERE code=?",(code,)); conn.commit()
            await message.reply_text("✅ Промокод выключен." if cur.rowcount else "❌ Код не найден.")
            return
        if low in ("!promo","!промо"):
            if len(parts)<3:
                await message.reply_text("Форматы:\n!promo SERTAL10 10%\n!promo BONUS2000 2000\nМожно: !promo CODE 10% 5000 100")
                return
            code=parts[1].upper(); raw=parts[2].replace(",",".")
            kind="percent" if raw.endswith("%") else "fixed"
            try: value=float(raw.rstrip("%"))
            except: await message.reply_text("❌ Скидка должна быть числом, например 10% или 2000"); return
            min_order=0.0; max_uses=None
            if len(parts)>=4:
                try:min_order=float(parts[3].replace(",","."))
                except: pass
            if len(parts)>=5:
                try:max_uses=int(parts[4])
                except: pass
            if kind=="percent" and not (0<value<=100): await message.reply_text("❌ Процент: от 1 до 100"); return
            if kind=="fixed" and value<=0: await message.reply_text("❌ Сумма скидки должна быть больше 0"); return
            conn.execute("INSERT INTO promos(code,kind,value,min_order,max_uses,used_count,active,created_at) VALUES(?,?,?,?,?,?,1,?) ON CONFLICT(code) DO UPDATE SET kind=excluded.kind,value=excluded.value,min_order=excluded.min_order,max_uses=excluded.max_uses,active=1",(code,kind,value,min_order,max_uses,0,int(time.time())))
            conn.commit()
            await message.reply_text(f"✅ Промокод {code} создан/обновлён. Скидка: {'%g%%'%value if kind=='percent' else '%g ֏'%value}. Мин.: {min_order:.0f} ֏. Лимит: {max_uses if max_uses is not None else 'без лимита'}.")
            return
    finally:
        conn.close()


async def admin_reply_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    message = update.message
    if not message or not message.text or not update.effective_user:
        return
    admin_id = update.effective_user.id
    if admin_id not in ADMIN_IDS:
        return
    text = message.text.strip()
    chat_type = message.chat.type if message.chat else "private"
    low = text.lower()

    if low in ("!help","!помощь"):
        await message.reply_text("SERTAL DELIVERY — команды админа\n\n!logsfile — база клиентов\n!excel — заказы Excel\n!ai вопрос — ИИ + интернет-поиск\n!promo CODE 10% — процентная скидка\n!promo CODE 2000 — скидка 2000 ֏\n!promo CODE 10% 5000 100 — 10%, мин. 5000 ֏, до 100 использований\n!promos — список промокодов\n!delpromo CODE — выключить промокод\n!supportgroup — назначить эту группу поддержкой\n!чек #123 — красивый чек реального заказа")
        return

    if low.startswith("!promo") or low.startswith("!promos") or low.startswith("!delpromo") or low.startswith("!deletepromo") or low.startswith("!промо") or low.startswith("!промокоды"):
        await promo_group_command(update, context)
        return

    if low == "!supportgroup":
        if chat_type == "private":
            await message.reply_text("Эту команду нужно выполнить внутри рабочей Telegram-группы поддержки.")
        else:
            set_support_group_chat_id(message.chat.id)
            await message.reply_text(f"✅ Эта группа назначена рабочей поддержкой SERTAL DELIVERY.\nID группы: {message.chat.id}")
        return

    # Explicit admin tools are allowed in groups and private chats.
    if low == "!logsfile":
        path = export_customers_xlsx()
        image = make_customers_report_image()
        try:
            with open(image, "rb") as fh:
                await message.reply_photo(photo=fh, caption="SERTAL DELIVERY — клиентская база")
            with open(path, "rb") as fh:
                await message.reply_document(document=fh, filename=path.name, caption="Полная база клиентов Excel")
        except Exception as e:
            await message.reply_text(f"Ошибка выгрузки: {e}")
        return

    if low == "!excel":
        path = export_orders_xlsx()
        image = make_orders_report_image()
        try:
            with open(image, "rb") as fh:
                await message.reply_photo(photo=fh, caption="SERTAL DELIVERY — таблица заказов")
            with open(path, "rb") as fh:
                await message.reply_document(document=fh, filename=path.name, caption="Полная таблица заказов Excel")
        except Exception as e:
            await message.reply_text(f"Ошибка Excel: {e}")
        return

    if low.startswith("!ai"):
        question=text[3:].strip()
        answer=await ai_group_answer(question)
        await message.reply_text(f"SERTAL DELIVERY · AI\n\n{answer}")
        return

    # Real order receipt: !чек #123 / !чек 123
    if low.startswith("!чек") or low.startswith("!check") or low.startswith("!receipt"):
        match = re.search(r"#?(\d+)", text)
        if not match:
            await message.reply_text("Формат: !чек #123")
            return
        order_id = int(match.group(1))
        conn = db()
        order = conn.execute("SELECT id,customer_id,status FROM orders WHERE id=?", (order_id,)).fetchone()
        conn.close()
        if not order:
            await message.reply_text(f"❌ Заказ №{order_id} не найден.")
            return
        path = make_order_receipt_image(order_id)
        if not path:
            await message.reply_text("❌ Не удалось сформировать чек.")
            return
        try:
            with open(path, "rb") as fh:
                await message.reply_photo(photo=fh, caption=f"SERTAL DELIVERY · Чек заказа №{order_id}")
        except Exception as e:
            await message.reply_text(f"Ошибка отправки чека: {e}")
        return

    # Support replies work both in the dedicated support group and in private admin chat.
    reply = message.reply_to_message
    if not reply:
        return
    bridge_chat_id = message.chat.id if chat_type != "private" else admin_id
    support_group_id = get_support_group_chat_id()
    if chat_type != "private" and support_group_id and bridge_chat_id != support_group_id:
        return
    conn = db()
    bridge = conn.execute("SELECT user_id FROM admin_message_bridge WHERE admin_telegram_id=? AND telegram_message_id=?", (bridge_chat_id, reply.message_id)).fetchone()
    conn.close()
    if not bridge:
        return
    add_chat_message(bridge["user_id"], "admin", text)
    conn = db()
    target = conn.execute("SELECT telegram_id FROM users WHERE id=?", (bridge["user_id"],)).fetchone()
    conn.close()
    if target and target["telegram_id"] and telegram_app:
        try:
            await telegram_app.bot.send_message(target["telegram_id"], f"SERTAL DELIVERY\n\nОтвет поддержки:\n{text}")
        except Exception as exc:
            print("SEND CLIENT:", exc)


async def location_handler(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    if not update.message.location:
        return

    telegram_id = update.effective_user.id

    lat = update.message.location.latitude
    lon = update.message.location.longitude

    conn = db()

    user = conn.execute("""
        SELECT id
        FROM users
        WHERE telegram_id=?
        AND role='courier'
        AND active=1
    """, (telegram_id,)).fetchone()

    if user:

        conn.execute("""
            UPDATE couriers
            SET lat=?,lon=?,updated_at=?
            WHERE user_id=?
        """, (
            lat,
            lon,
            int(time.time()),
            user["id"]
        ))

        conn.commit()

    conn.close()


# =========================================================
# CLEANUP
# =========================================================

async def notify_deploy_started():
    """Notify the configured support group after Render starts the new release."""
    if not telegram_app:
        return
    chat_id = get_support_group_chat_id()
    if not chat_id:
        return
    commit = os.getenv("RENDER_GIT_COMMIT", "").strip()
    commit_short = commit[:7] if commit else "новая версия"
    service = os.getenv("RENDER_SERVICE_NAME", "SERTAL DELIVERY")
    message = (
        "🤖 Мой раб обновил меня.\n\n"
        "🚀 Обнова запущена и Render поднял новую версию.\n"
        f"📦 Сервис: {service}\n"
        f"🔖 Версия: {commit_short}\n"
        "✅ SERTAL DELIVERY снова в строю."
    )
    try:
        await telegram_app.bot.send_message(chat_id=chat_id, text=message)
    except Exception as exc:
        print("Deploy notification:", exc)

async def cleanup_loop():

    while True:

        try:
            cleanup_old_closed()
        except Exception as e:
            print("Cleanup:", e)

        await asyncio.sleep(30)


@app.on_event("startup")
async def startup():

    global telegram_app

    if not BOT_TOKEN:
        print("WARNING: BOT_TOKEN is missing")
        return

    telegram_app = (
        __import__("telegram.ext", fromlist=["Application"])
        .Application.builder()
        .token(BOT_TOKEN)
        .build()
    )

    telegram_app.add_handler(
        CommandHandler("start", start_command)
    )

    telegram_app.add_handler(
        CommandHandler("logsfile", lambda update, context: logsfile_direct(update, context))
    )

    telegram_app.add_handler(
        MessageHandler(
            filters.CONTACT,
            contact_handler
        )
    )

    telegram_app.add_handler(
        MessageHandler(
            filters.LOCATION,
            location_handler
        )
    )

    telegram_app.add_handler(
        MessageHandler(
            filters.TEXT & ~filters.COMMAND,
            admin_reply_handler
        )
    )

    await telegram_app.initialize()
    await telegram_app.start()
    await telegram_app.updater.start_polling()

    await notify_deploy_started()

    asyncio.create_task(
        cleanup_loop()
    )

    print("SERTAL DELIVERY started")


@app.on_event("shutdown")
async def shutdown():

    global telegram_app

    if telegram_app:

        try:
            await telegram_app.updater.stop()
            await telegram_app.stop()
            await telegram_app.shutdown()
        except Exception as e:
            print("Telegram shutdown:", e)


# =========================================================
# RUN
# =========================================================

if __name__ == "__main__":
    uvicorn.run(
        app,
        host="0.0.0.0",
        port=int(os.getenv("PORT", "10000"))
    )
